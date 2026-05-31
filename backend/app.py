import os
import uuid
import secrets
import hashlib
import io
import json
import csv
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

from flask import Flask, jsonify, request, session, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect as sa_inspect, text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

DB_USER = os.getenv("DB_USER", "jkh")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_HOST = os.getenv("DB_HOST", "mysql")
DB_PORT = os.getenv("DB_PORT", "3306")
DB_NAME = os.getenv("DB_NAME", "jkh")

app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "change-me-please")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.getenv("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = os.getenv("SESSION_COOKIE_SECURE", "0") == "1"
app.config["IMPORT_MAX_UPLOAD_BYTES"] = int(os.getenv("IMPORT_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
app.config["IMPORT_UPLOAD_BLOB_TTL_DAYS"] = int(os.getenv("IMPORT_UPLOAD_BLOB_TTL_DAYS", "14"))

db = SQLAlchemy(app)


def sqlite_autoincrement_bigint_pk():
    return db.BigInteger().with_variant(db.Integer, "sqlite")

IMPORT_BATCH_CRITICAL_COLUMNS = (
    "rows_skipped",
    "file_name",
    "uploaded_by",
    "error_message",
)


ABONENT_SUMMARY_DIRTY_REASONS = {
    "PAYMENTS_CHANGED",
    "IMPORT_PAYMENTS_APPLIED",
    "CALC_PERIOD_CHANGED",
    "EXCLUDES_CHANGED",
    "MORATORIUM_CHANGED",
    "RESPONSIBILITY_CHANGED",
    "TARIFFS_CHANGED",
    "AUTOACCRUAL_CHANGED",
    "LEDGER_WRITE",
    "UNKNOWN_CHANGE",
}

AUDIT_REASON_DEFAULTS = {
    "summary": {
        "dirty": "INPUT_HASH_CHANGED",
        "missing": "SUMMARY_NOT_BUILT",
        "error": "CALC_ENGINE_UNAVAILABLE",
        "invalid": "SUMMARY_JSON_INVALID",
    },
    "snapshot": {
        "dirty": "INPUT_HASH_CHANGED",
        "missing": "CARD_SNAPSHOT_MISSING",
        "error": "CALC_ENGINE_UNAVAILABLE",
        "invalid": "CARD_SNAPSHOT_JSON_INVALID",
    },
}

IMPORT_BATCH_AUDIT_FIELDS_MIGRATION_SQL = (
    "ALTER TABLE import_batches "
    "ADD COLUMN rows_skipped INT NOT NULL DEFAULT 0, "
    "ADD COLUMN file_name VARCHAR(255) NULL, "
    "ADD COLUMN uploaded_by VARCHAR(255) NULL, "
    "ADD COLUMN error_message TEXT NULL;"
)



RECALC_BATCH_ACTIVE_STATUSES = {"queued", "running"}
RECALC_BATCH_FINAL_STATUSES = {"completed", "done", "failed", "stale"}
RECALC_BATCH_RUNNING_TTL_SECONDS = 30 * 60
RECALC_BATCH_MAX_UIDS = 100
RECALC_BATCH_KEEP_PER_OWNER = 20
RECALC_BATCH_RETENTION_DAYS = 7
SNAPSHOT_ENGINE_VERSION = "JKHCalcEngine:stage16-mvp"

class User(db.Model):
    __tablename__ = "users"

    id = db.Column(db.String(64), primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(32), nullable=False, default="user")
    display_name = db.Column(db.String(255), nullable=False, default="")
    disabled = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    last_login = db.Column(db.DateTime, nullable=True)


class KVStore(db.Model):
    __tablename__ = "kv_store"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    owner = db.Column(db.String(128), nullable=False)
    k = db.Column(db.String(255), nullable=False)
    v = db.Column(db.Text, nullable=False)
    updated_at = db.Column(
        db.DateTime,
        server_default=text("CURRENT_TIMESTAMP"),
        onupdate=text("CURRENT_TIMESTAMP"),
    )

    __table_args__ = (db.UniqueConstraint("owner", "k", name="uq_owner_key"),)


class ImportBatch(db.Model):
    __tablename__ = "import_batches"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    created_by_user_id = db.Column(db.String(64), nullable=False, index=True)

    original_filename = db.Column(db.String(255), nullable=False)
    file_sha256 = db.Column(db.String(64), nullable=False, index=True)
    upload_blob = db.Column(db.LargeBinary, nullable=False)

    display_name = db.Column(db.String(255), nullable=False, default="")
    accounting_year = db.Column(db.String(16), nullable=False, default="")
    version_label = db.Column(db.String(64), nullable=False, default="")
    notes = db.Column(db.Text, nullable=False, default="")

    duplicate_of_batch_id = db.Column(db.Integer, nullable=True)
    supersedes_batch_id = db.Column(db.Integer, nullable=True)
    overlaps_with_batch_id = db.Column(db.Integer, nullable=True)

    is_exact_duplicate = db.Column(db.Boolean, nullable=False, default=False)
    has_period_overlap = db.Column(db.Boolean, nullable=False, default=False)

    rows_total = db.Column(db.Integer, nullable=False, default=0)
    rows_valid = db.Column(db.Integer, nullable=False, default=0)
    rows_invalid = db.Column(db.Integer, nullable=False, default=0)
    rows_duplicate = db.Column(db.Integer, nullable=False, default=0)
    rows_applied = db.Column(db.Integer, nullable=False, default=0)
    rows_skipped = db.Column(db.Integer, nullable=False, default=0)

    file_name = db.Column(db.String(255), nullable=False, default="")
    uploaded_by = db.Column(db.String(64), nullable=False, default="")
    error_message = db.Column(db.Text, nullable=False, default="")

    status = db.Column(db.String(32), nullable=False, default="uploaded", index=True)

    uploaded_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)


class ImportBatchRow(db.Model):
    __tablename__ = "import_rows"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    batch_id = db.Column(db.Integer, db.ForeignKey("import_batches.id"), nullable=False, index=True)
    row_no = db.Column(db.Integer, nullable=False)
    excel_sheet_name = db.Column(db.String(255), nullable=False, default="")
    excel_row_ref = db.Column(db.String(64), nullable=False, default="")

    raw_payload_json = db.Column(db.Text, nullable=False, default="{}")
    normalized_payload_json = db.Column(db.Text, nullable=False, default="{}")

    account_uid = db.Column(db.String(128), nullable=False, default="", index=True)
    abonent_id = db.Column(db.String(128), nullable=False, default="")
    account_number = db.Column(db.String(128), nullable=False, default="")

    payment_date = db.Column(db.String(10), nullable=False, default="")
    payment_period = db.Column(db.String(7), nullable=False, default="")
    charge_year = db.Column(db.String(4), nullable=False, default="")
    charge_month = db.Column(db.String(2), nullable=False, default="")
    amount = db.Column(db.String(32), nullable=False, default="")
    paid_date = db.Column(db.String(10), nullable=False, default="")

    source_index = db.Column(db.Integer, nullable=True)
    source_label = db.Column(db.String(255), nullable=False, default="")

    fingerprint = db.Column(db.String(64), nullable=False, default="", index=True)
    matched_payment_id = db.Column(db.String(64), nullable=False, default="")
    applied_at = db.Column(db.DateTime, nullable=True)

    status = db.Column(db.String(32), nullable=False, default="parsed", index=True)
    reason_code = db.Column(db.String(64), nullable=False, default="")
    reason_text = db.Column(db.String(1024), nullable=False, default="")
    error_details_json = db.Column(db.Text, nullable=False, default="{}")


class ImportAppliedFingerprint(db.Model):
    __tablename__ = "import_applied_fingerprints"

    id = db.Column(sqlite_autoincrement_bigint_pk(), primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(191), nullable=False, index=True)
    import_type = db.Column(db.String(32), nullable=False, default="payments")
    fingerprint = db.Column(db.String(255), nullable=False)
    account_uid = db.Column(db.String(191), nullable=True)
    account_number = db.Column(db.String(191), nullable=True)
    payment_period = db.Column(db.String(7), nullable=True)
    paid_date = db.Column(db.Date, nullable=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    source_index = db.Column(db.Integer, nullable=False, default=1)
    payment_id = db.Column(db.String(64), nullable=False, default="")
    batch_id = db.Column(db.BigInteger, nullable=True, index=True)
    created_at = db.Column(db.DateTime, nullable=False, server_default=text("CURRENT_TIMESTAMP"))

    __table_args__ = (
        db.UniqueConstraint("owner_id", "import_type", "fingerprint", name="uq_owner_import_fp"),
    )


class PaymentAuditLog(db.Model):
    __tablename__ = "payment_audit_log"

    id = db.Column(sqlite_autoincrement_bigint_pk(), primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    batch_id = db.Column(db.Integer, nullable=False, index=True)
    row_id = db.Column(db.Integer, nullable=True, index=True)
    action = db.Column(db.String(32), nullable=False)
    status = db.Column(db.String(32), nullable=False)
    details_json = db.Column(db.Text, nullable=False, default="{}")
    created_at = db.Column(db.DateTime, nullable=False, server_default=text("CURRENT_TIMESTAMP"))


class AbonentSummary(db.Model):
    __tablename__ = "abonent_summary"

    id = db.Column(sqlite_autoincrement_bigint_pk(), primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    abonent_id = db.Column(db.String(128), nullable=False, default="", index=True)
    account_uid = db.Column(db.String(128), nullable=False, default="", index=True)
    account_number = db.Column(db.String(128), nullable=False, default="", index=True)
    fio = db.Column(db.String(255), nullable=False, default="")
    address = db.Column(db.String(1024), nullable=False, default="")
    total_accrued = db.Column(db.Numeric(14, 2), nullable=True)
    total_paid = db.Column(db.Numeric(14, 2), nullable=True)
    main_debt = db.Column(db.Numeric(14, 2), nullable=True)
    penalty_debt = db.Column(db.Numeric(14, 2), nullable=True)
    total_debt = db.Column(db.Numeric(14, 2), nullable=True)
    summary_status = db.Column(db.String(32), nullable=False, default="missing", index=True)
    summary_reason = db.Column(db.String(128), nullable=False, default="")
    input_hash = db.Column(db.String(64), nullable=False, default="", index=True)
    dirty_since = db.Column(db.DateTime, nullable=True)
    last_error_code = db.Column(db.String(64), nullable=False, default="")
    summary_json = db.Column(db.Text, nullable=False, default="{}")
    created_at = db.Column(db.DateTime, nullable=False, server_default=text("CURRENT_TIMESTAMP"))
    updated_at = db.Column(
        db.DateTime,
        server_default=text("CURRENT_TIMESTAMP"),
        onupdate=text("CURRENT_TIMESTAMP"),
    )


class CardSnapshot(db.Model):
    __tablename__ = "card_snapshot"

    id = db.Column(sqlite_autoincrement_bigint_pk(), primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    abonent_uid = db.Column(db.String(128), nullable=False, default="", index=True)
    abonent_id = db.Column(db.String(128), nullable=False, default="", index=True)
    snapshot_status = db.Column(db.String(32), nullable=False, default="missing", index=True)
    snapshot_reason = db.Column(db.String(128), nullable=False, default="")
    input_hash = db.Column(db.String(64), nullable=False, default="", index=True)
    ledger_version = db.Column(db.String(64), nullable=False, default="")
    tariff_version = db.Column(db.String(64), nullable=False, default="")
    rate_version = db.Column(db.String(64), nullable=False, default="")
    exclude_version = db.Column(db.String(64), nullable=False, default="")
    links_version = db.Column(db.String(64), nullable=False, default="")
    engine_version = db.Column(db.String(64), nullable=False, default=SNAPSHOT_ENGINE_VERSION)
    computed_at = db.Column(db.DateTime, nullable=True)
    updated_at = db.Column(
        db.DateTime,
        server_default=text("CURRENT_TIMESTAMP"),
        onupdate=text("CURRENT_TIMESTAMP"),
    )
    snapshot_json = db.Column(db.Text, nullable=False, default="{}")

    __table_args__ = (db.UniqueConstraint("owner_id", "abonent_uid", name="uq_card_snapshot_owner_uid"),)


class RecalcUidLock(db.Model):
    __tablename__ = "recalc_uid_locks"

    id = db.Column(sqlite_autoincrement_bigint_pk(), primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    abonent_uid = db.Column(db.String(128), nullable=False, index=True)
    lock_token = db.Column(db.String(64), nullable=False, default="")
    status = db.Column(db.String(32), nullable=False, default="running", index=True)
    started_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime,
        server_default=text("CURRENT_TIMESTAMP"),
        onupdate=text("CURRENT_TIMESTAMP"),
    )

    __table_args__ = (db.UniqueConstraint("owner_id", "abonent_uid", name="uq_recalc_uid_lock_owner_uid"),)


class RecalcBatchJob(db.Model):
    __tablename__ = "recalc_batch_jobs"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    requested_by = db.Column(db.String(64), nullable=False, default="")
    reason = db.Column(db.String(64), nullable=False, default="")
    status = db.Column(db.String(32), nullable=False, default="queued", index=True)
    total_count = db.Column(db.Integer, nullable=False, default=0)
    processed_count = db.Column(db.Integer, nullable=False, default=0)
    fresh_count = db.Column(db.Integer, nullable=False, default=0)
    error_count = db.Column(db.Integer, nullable=False, default=0)
    skipped_count = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, server_default=text("CURRENT_TIMESTAMP"))
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)
    error_message = db.Column(db.Text, nullable=False, default="")


class RecalcBatchJobItem(db.Model):
    __tablename__ = "recalc_batch_job_items"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    job_id = db.Column(db.Integer, db.ForeignKey("recalc_batch_jobs.id"), nullable=False, index=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    account_uid = db.Column(db.String(128), nullable=False, default="", index=True)
    status = db.Column(db.String(32), nullable=False, default="queued", index=True)
    summary_status = db.Column(db.String(32), nullable=False, default="")
    summary_reason = db.Column(db.String(128), nullable=False, default="")
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)
    error_message = db.Column(db.Text, nullable=False, default="")


class BulkCalcVerifyJob(db.Model):
    __tablename__ = "bulk_calc_verify_jobs"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    requested_by = db.Column(db.String(64), nullable=False, default="")
    reason = db.Column(db.String(64), nullable=False, default="")
    status = db.Column(db.String(32), nullable=False, default="queued", index=True)
    total_count = db.Column(db.Integer, nullable=False, default=0)
    processed_count = db.Column(db.Integer, nullable=False, default=0)
    ok_count = db.Column(db.Integer, nullable=False, default=0)
    mismatch_count = db.Column(db.Integer, nullable=False, default=0)
    error_count = db.Column(db.Integer, nullable=False, default=0)
    skipped_count = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, server_default=text("CURRENT_TIMESTAMP"))
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)
    error_message = db.Column(db.Text, nullable=False, default="")


class BulkCalcVerifyJobItem(db.Model):
    __tablename__ = "bulk_calc_verify_job_items"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    job_id = db.Column(db.Integer, db.ForeignKey("bulk_calc_verify_jobs.id"), nullable=False, index=True)
    owner_id = db.Column(db.String(128), nullable=False, index=True)
    account_uid = db.Column(db.String(128), nullable=False, default="", index=True)
    status = db.Column(db.String(32), nullable=False, default="queued", index=True)
    reason = db.Column(db.String(128), nullable=False, default="")
    old_summary_json = db.Column(db.Text, nullable=False, default="{}")
    new_summary_json = db.Column(db.Text, nullable=False, default="{}")
    diff_json = db.Column(db.Text, nullable=False, default="{}")
    error_code = db.Column(db.String(64), nullable=False, default="")
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)
    error_message = db.Column(db.Text, nullable=False, default="")


def _json_error(error: str, code: int):
    return jsonify(ok=False, error=error), code


def _parse_pagination_args(default_per_page: int = 50, max_per_page: int = 200):
    try:
        page = int(request.args.get("page", "1"))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get("per_page", request.args.get("limit", str(default_per_page))))
    except (TypeError, ValueError):
        per_page = default_per_page

    page = max(1, page)
    per_page = max(1, min(max_per_page, per_page))
    return page, per_page


def _pagination_payload(page: int, per_page: int, total: int):
    pages = (total + per_page - 1) // per_page if per_page else 0
    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "pages": pages,
        "has_next": page < pages,
        "has_prev": page > 1,
    }


def _table_columns(table_name: str):
    try:
        inspector = sa_inspect(db.engine)
        if not inspector.has_table(table_name):
            return set()
        return {col["name"] for col in inspector.get_columns(table_name)}
    except SQLAlchemyError:
        db.session.rollback()
        raise


def _first_existing_column(columns, candidates):
    for name in candidates:
        if name in columns:
            return name
    return None


def _sql_ident(name: str) -> str:
    return "`" + str(name).replace("`", "``") + "`"


def _sql_cast_text(expr: str) -> str:
    if db.engine.dialect.name == "mysql":
        return f"CAST({expr} AS CHAR)"
    return f"CAST({expr} AS TEXT)"


def _decimal_json_or_none(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return str(value)
    return value


def _dt_json_or_none(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat() + "Z"
    return str(value)


def _cache_status(value: str, default: str = "missing"):
    status = _norm_text(value).lower()
    if status in {"fresh", "dirty", "missing", "error", "invalid"}:
        return status
    return default


def _cache_reason(kind: str, status: str, reason: str):
    clean_status = _cache_status(status)
    clean_reason = _norm_text(reason)
    if clean_status == "fresh":
        return clean_reason
    if clean_reason:
        return clean_reason
    return AUDIT_REASON_DEFAULTS.get(kind, {}).get(clean_status, "UNKNOWN_REASON")


def _abonent_summary_payload(row: AbonentSummary):
    summary, _parse_error = _safe_summary_from_json_for_columns(row.summary_json)

    return {
        "id": row.id,
        "owner_id": row.owner_id,
        "abonent_id": row.abonent_id or "",
        "account_uid": row.account_uid or "",
        "account_number": row.account_number or "",
        "summary": summary,
        "created_at": row.created_at.isoformat() + "Z" if row.created_at else None,
        "updated_at": row.updated_at.isoformat() + "Z" if row.updated_at else None,
    }


def _decimal_or_none(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        d = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return None
    if not d.is_finite():
        return None
    return d


def _summary_value(summary: dict, *keys):
    totals = summary.get("totals") if isinstance(summary.get("totals"), dict) else {}
    for key in keys:
        if key.startswith("totals."):
            value = totals.get(key.split(".", 1)[1])
        else:
            value = summary.get(key)
        if value is not None and not (isinstance(value, str) and not value.strip()):
            return value
    return None


def _apply_abonent_summary_columns(row: AbonentSummary, target: dict, summary: dict):
    identity = target.get("identity") if isinstance(target, dict) and isinstance(target.get("identity"), dict) else {}
    summary = summary if isinstance(summary, dict) else {}
    abonent = summary.get("abonent") if isinstance(summary.get("abonent"), dict) else {}
    status = _summary_column_status_from_payload(summary)
    reason = _cache_reason("summary", status, summary.get("summary_reason") or summary.get("reason"))

    row.fio = _norm_text(summary.get("fio") or abonent.get("fio") or identity.get("fio"))
    row.address = _norm_text(summary.get("address") or abonent.get("address") or identity.get("address"))
    row.total_accrued = _decimal_or_none(_summary_value(summary, "total_accrued", "totals.total_accrued", "totals.accrued"))
    row.total_paid = _decimal_or_none(_summary_value(summary, "total_paid", "totals.total_paid", "totals.paid"))
    row.main_debt = _decimal_or_none(_summary_value(summary, "main_debt", "principal", "total_principal", "totals.principal"))
    row.penalty_debt = _decimal_or_none(_summary_value(summary, "penalty_debt", "penalty", "total_penalty", "totals.penalty", "totals.total_penalty"))
    row.total_debt = _decimal_or_none(_summary_value(summary, "total_debt", "total", "totals.total_debt", "totals.total", "totals.debt"))
    row.summary_status = status
    row.summary_reason = reason
    row.input_hash = _norm_text(summary.get("input_hash"))
    row.last_error_code = reason if status in {"error", "invalid"} else ""
    if status == "dirty" and not row.dirty_since:
        row.dirty_since = datetime.utcnow()
    elif status == "fresh":
        row.dirty_since = None


def _set_abonent_summary_json(row: AbonentSummary, target: dict, summary: dict):
    row.summary_json = json.dumps(summary if isinstance(summary, dict) else {}, ensure_ascii=False, sort_keys=True)
    _apply_abonent_summary_columns(row, target, summary)


def _safe_summary_from_json_for_columns(raw_json: str):
    try:
        payload = json.loads(raw_json or "{}")
    except (TypeError, ValueError):
        return {"summary_status": "invalid", "summary_reason": "SUMMARY_JSON_INVALID"}, "SUMMARY_JSON_INVALID"
    if not isinstance(payload, dict):
        return {"summary_status": "invalid", "summary_reason": "SUMMARY_JSON_NOT_OBJECT"}, "SUMMARY_JSON_NOT_OBJECT"
    return payload, ""


def _summary_column_status_from_payload(summary: dict | None):
    if not isinstance(summary, dict):
        return "invalid"
    status = _norm_text(summary.get("summary_status") or summary.get("status")).lower()
    if status in {"fresh", "dirty", "missing", "error", "invalid"}:
        return status
    return "missing"


def _column_decimal_equal(left, right):
    left_dec = _decimal_or_none(left)
    right_dec = _decimal_or_none(right)
    return left_dec == right_dec


def _expected_abonent_summary_column_row(row: AbonentSummary):
    summary, parse_error = _safe_summary_from_json_for_columns(row.summary_json)
    expected = AbonentSummary(
        owner_id=row.owner_id,
        abonent_id=row.abonent_id or "",
        account_uid=row.account_uid or "",
        account_number=row.account_number or "",
    )
    _apply_abonent_summary_columns(expected, {}, summary)
    if parse_error:
        expected.last_error_code = parse_error
    return expected, parse_error


def _abonent_summary_consistency_mismatches(row: AbonentSummary):
    expected, parse_error = _expected_abonent_summary_column_row(row)
    checks = (
        ("summary_status", _norm_text(row.summary_status), _norm_text(expected.summary_status)),
        ("summary_reason", _norm_text(row.summary_reason), _norm_text(expected.summary_reason)),
        ("total_debt", row.total_debt, expected.total_debt),
        ("total_accrued", row.total_accrued, expected.total_accrued),
        ("total_paid", row.total_paid, expected.total_paid),
        ("penalty_debt", row.penalty_debt, expected.penalty_debt),
    )
    mismatches = []
    for field, current, expected_value in checks:
        if field.startswith("total_") or field == "penalty_debt":
            if not _column_decimal_equal(current, expected_value):
                mismatches.append({"field": field, "column": current, "expected": expected_value})
        elif current != expected_value:
            mismatches.append({"field": field, "column": current, "expected": expected_value})
    return mismatches, expected, parse_error


def audit_abonent_summary_consistency(apply: bool = False, sample_limit: int = 20, owner_id: str = ""):
    query = AbonentSummary.query
    if _norm_text(owner_id):
        query = query.filter_by(owner_id=_norm_text(owner_id))

    result = {
        "checked": 0,
        "mismatch_count": 0,
        "updated": 0,
        "samples": [],
    }
    for row in query.order_by(AbonentSummary.id.asc()).all():
        result["checked"] += 1
        mismatches, expected, parse_error = _abonent_summary_consistency_mismatches(row)
        if not mismatches:
            continue
        result["mismatch_count"] += 1
        if len(result["samples"]) < sample_limit:
            result["samples"].append({
                "id": row.id,
                "owner_id": row.owner_id,
                "account_uid": row.account_uid,
                "account_number": row.account_number,
                "parse_error": parse_error,
                "mismatches": [
                    {
                        "field": item["field"],
                        "column": str(item["column"]) if item["column"] is not None else None,
                        "expected": str(item["expected"]) if item["expected"] is not None else None,
                    }
                    for item in mismatches
                ],
            })
        if apply:
            row.total_accrued = expected.total_accrued
            row.total_paid = expected.total_paid
            row.penalty_debt = expected.penalty_debt
            row.total_debt = expected.total_debt
            row.summary_status = expected.summary_status
            row.summary_reason = expected.summary_reason
            row.last_error_code = expected.last_error_code
            row.dirty_since = expected.dirty_since
            result["updated"] += 1

    if apply and result["updated"]:
        db.session.commit()
    return result


def _normalize_email(value: str) -> str:
    return str(value or "").strip().lower()


def _user_payload(user: User):
    return {
        "id": user.id,
        "email": user.email,
        "role": user.role,
        "displayName": user.display_name or "",
        "disabled": bool(user.disabled),
        "createdAt": int(user.created_at.timestamp() * 1000) if user.created_at else 0,
        "lastLogin": int(user.last_login.timestamp() * 1000) if user.last_login else 0,
    }


def _current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    user = User.query.filter_by(id=user_id).first()
    if not user or user.disabled:
        session.clear()
        return None
    return user


def _require_user():
    user = _current_user()
    if not user:
        return None, _json_error("not_authenticated", 401)
    return user, None


def _require_admin():
    user, err = _require_user()
    if err:
        return None, err
    if user.role != "admin":
        return None, _json_error("forbidden", 403)
    return user, None


def _resolve_owner(explicit_owner: str | None = None, allow_admin_override: bool = False):
    user, err = _require_user()
    if err:
        return None, err
    wanted = str(explicit_owner or "").strip()
    if allow_admin_override and user.role == "admin" and wanted:
        return wanted, None
    return user.id, None


def _sync_log(action: str, owner: str, **extra):
    parts = [f"action={action}", f"owner={owner}"]
    for k, v in extra.items():
        parts.append(f"{k}={v}")
    app.logger.info("[JKH sync] %s", " ".join(parts))


GLOBAL_OWNER = "GLOBAL"
GLOBAL_KEYS = {
    "refinancing_rates_normal_v1",
    "refinancing_rates_moratorium_v1",
}

PROTECTED_OWNER_LEVEL_KEYS = {
    "tariffs_dynamic_v1",  # legacy read-only / migration only / excluded from upload
    "tariffs_content_repair_v1",  # legacy read-only / migration only / excluded from upload
    "tariffs_content_repair_v1_backup",  # legacy read-only / migration only / excluded from upload
}
PROTECTED_OWNER_LEVEL_PREFIXES = (
    "tariffs_",
    "ref_rates_",
)

IMPORT_BATCH_STATUSES = {
    "uploaded",
    "parsed",
    "validated",
    "ready_to_apply",
    "applying",
    "applied",
    "failed",
    "cancelled",
}

ROW_STATUSES = {
    "parsed",
    "validated",
    "invalid",
    "duplicate",
    "conflict",
    "ready",
    "applied",
    "skipped",
    "failed",
}

IMPORT_ALLOWED_TRANSITIONS = {
    "parse": {"uploaded", "failed"},
    "validate": {"parsed", "validated"},
    "apply": {"ready_to_apply"},
}

IMPORT_REQUIRED_COLUMNS = {"account_uid", "account_number", "payment_date", "payment_period", "amount", "source_index"}

HEADER_ALIASES = {
    "account_uid": {"account_uid", "uid", "уид", "лс uid", "лицевой uid"},
    "abonent_id": {"abonent_id", "абонент", "abonent"},
    "account_number": {"account_number", "лицевой счет", "лицевой счёт", "лс", "лицевой_счет"},
    "payment_date": {"payment_date", "paid_date", "дата", "дата платежа", "дата оплаты"},
    "payment_period": {"payment_period", "period", "период", "расчетный период", "месяц"},
    "amount": {"amount", "sum", "сумма", "сумма оплаты", "оплата"},
    "source_index": {"source_index", "source", "источник", "источник платежа", "индекс источника"},
}


def _is_protected_owner_level_key(base_key: str) -> bool:
    k = str(base_key or "").strip()
    if not k:
        return False
    if k in PROTECTED_OWNER_LEVEL_KEYS:
        return True
    return any(k.startswith(pref) for pref in PROTECTED_OWNER_LEVEL_PREFIXES)


def _effective_owner_for_key(owner: str, base_key: str) -> str:
    key = str(base_key or "").strip()
    if key in GLOBAL_KEYS:
        return GLOBAL_OWNER
    return owner


def _check_store_write_access(user: User, owner: str, base_key: str):
    key = str(base_key or "").strip()
    target_owner = str(owner or "").strip()

    if key in GLOBAL_KEYS and user.role != "admin":
        return False, "global_admin_only"

    if _is_protected_owner_level_key(key):
        if key.startswith("tariffs_") and user.role == "user":
            expected_key = f"tariffs_{target_owner}"
            if target_owner == user.id and key == expected_key:
                return True, "ok"
            return False, "tariffs_owner_mismatch"
        if user.role != "admin":
            return False, "protected_key_admin_only"

    return True, "ok"


def _require_write_access_for_key(owner: str, base_key: str):
    user, err = _require_user()
    if err:
        return err, None
    allowed, reason = _check_store_write_access(user, owner, base_key)
    if not allowed:
        if reason == "global_admin_only":
            return _json_error("global_admin_only", 403), reason
        return _json_error("forbidden", 403), reason
    return None, None


def _log_store_forbidden(user: User | None, owner: str, key: str, reason: str):
    app.logger.warning(
        "[store][forbidden] user_id=%s role=%s owner=%s key=%s reason=%s",
        getattr(user, "id", ""),
        getattr(user, "role", ""),
        owner,
        key,
        reason,
    )
    return None


def _norm_text(v):
    return str(v or "").strip()


def _norm_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return str(Decimal(str(v)).quantize(Decimal("0.01")))
        except InvalidOperation:
            return None
    s = _norm_text(v).replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return str(Decimal(s).quantize(Decimal("0.01")))
    except InvalidOperation:
        return None


def _norm_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = _norm_text(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _norm_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = _norm_text(v)
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None


def _norm_period(v):
    s = _norm_text(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})[-./](\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{1,2})[-./](\d{4})$", s)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    return None


def _norm_upload_payment_period(v):
    s = _norm_text(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})$", s)
    if not m:
        return None
    month = int(m.group(2))
    if month < 1 or month > 12:
        return None
    return s


def _raw_upload_payment_period_present(row):
    try:
        payload = json.loads(row.raw_payload_json or "{}")
    except Exception:
        payload = {}
    if isinstance(payload, dict) and "payment_period" in payload:
        return _norm_text(payload.get("payment_period")) != ""
    return False


def normalize_paid_date(v):
    value = _norm_date(v)
    if not value:
        raise ValueError("paid_date_invalid")
    return value


def normalize_payment_period(v):
    value = _norm_period(v)
    if not value:
        raise ValueError("payment_period_invalid")
    return value


def normalize_amount(v):
    value = _norm_amount(v)
    if not value:
        raise ValueError("amount_invalid")
    return value


def normalize_source_index(v):
    if v is None or _norm_text(v) == "":
        raise ValueError("source_index_required")
    try:
        idx = int(v)
    except Exception as ex:
        raise ValueError("source_index_invalid") from ex
    if idx < 1:
        raise ValueError("source_index_invalid")
    return idx


def normalize_account_number(v):
    value = _norm_text(v)
    if not value:
        raise ValueError("account_number_required")
    return value


def normalize_uid(v):
    value = _norm_text(v)
    if not value:
        raise ValueError("account_uid_required")
    return value


def to_ledger_paid_date(paid_date_iso: str) -> str:
    d = datetime.strptime(paid_date_iso, "%Y-%m-%d").date()
    return d.strftime("%d.%m.%Y")


def payment_fingerprint(account_uid, paid_date, amount, source_index):
    uid_norm = normalize_uid(account_uid)
    paid_date_norm = normalize_paid_date(paid_date)
    amount_norm = normalize_amount(amount)
    source_index_norm = normalize_source_index(source_index)
    raw = "|".join([
        uid_norm,
        paid_date_norm,
        amount_norm,
        str(source_index_norm),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def build_payment_fingerprint(owner_id, account_uid, account_number, paid_date, amount, source_index, payment_period):
    return payment_fingerprint(account_uid, paid_date, amount, source_index)



class LedgerJsonInvalidError(ValueError):
    pass


def _load_existing_payment_ledger_or_raise(kv):
    if not kv:
        return []
    raw = kv.v
    try:
        ledger = json.loads(raw)
    except Exception as exc:
        raise LedgerJsonInvalidError("LEDGER_JSON_INVALID") from exc
    if not isinstance(ledger, list):
        raise LedgerJsonInvalidError("LEDGER_JSON_INVALID")
    return ledger

def _classify_payment(account_uid, paid_date, amount, ledger_items):
    paid_date_norm = normalize_paid_date(paid_date)
    amount_norm = normalize_amount(amount)
    account_uid_norm = normalize_uid(account_uid)
    for item in ledger_items:
        item_date = _norm_date(item.get("paid_date"))
        if not item_date:
            continue
        if item_date != paid_date_norm:
            continue

        item_uid = _norm_text(item.get("uid"))
        if item_uid and item_uid != account_uid_norm:
            continue

        item_amount = _norm_amount(item.get("paid"))
        if item_amount == amount_norm:
            return "DUPLICATE"
        return "CONFLICT"
    return "NEW_PAYMENT"


def _classify_import_payment(account_uid, paid_date, amount, fingerprint, ledger_items):
    paid_date_norm = normalize_paid_date(paid_date)
    amount_norm = normalize_amount(amount)
    account_uid_norm = normalize_uid(account_uid)
    fingerprint_norm = _norm_text(fingerprint)
    for item in ledger_items:
        item_date = _norm_date(item.get("paid_date"))
        if not item_date or item_date != paid_date_norm:
            continue

        item_uid = _norm_text(item.get("uid"))
        if item_uid and item_uid != account_uid_norm:
            continue

        item_amount = _norm_amount(item.get("paid"))
        if item_amount is None:
            continue
        if item_amount == amount_norm:
            item_fingerprint = _norm_text(item.get("fingerprint"))
            if item_fingerprint:
                if item_fingerprint == fingerprint_norm:
                    return "DUPLICATE"
                continue
            return "DUPLICATE"
        return "CONFLICT"
    return "NEW_PAYMENT"


def _extract_year_month(period):
    if not period:
        return "", ""
    yy, mm = period.split("-", 1)
    return yy, mm


def _batch_payload(batch: ImportBatch):
    return {
        "id": batch.id,
        "owner_id": batch.owner_id,
        "created_by_user_id": batch.created_by_user_id,
        "display_name": batch.display_name,
        "original_filename": batch.original_filename,
        "accounting_year": batch.accounting_year,
        "version_label": batch.version_label,
        "notes": batch.notes,
        "status": batch.status,
        "is_exact_duplicate": bool(batch.is_exact_duplicate),
        "has_period_overlap": bool(batch.has_period_overlap),
        "rows_total": batch.rows_total,
        "rows_valid": batch.rows_valid,
        "rows_invalid": batch.rows_invalid,
        "rows_duplicate": batch.rows_duplicate,
        "rows_applied": batch.rows_applied,
        "rows_skipped": batch.rows_skipped,
        "file_name": batch.file_name,
        "uploaded_by": batch.uploaded_by,
        "error_message": batch.error_message,
        "uploaded_at": int(batch.uploaded_at.timestamp() * 1000) if batch.uploaded_at else 0,
    }


def _row_payload(r: ImportBatchRow):
    return {
        "id": r.id,
        "batch_id": r.batch_id,
        "row_no": r.row_no,
        "excel_sheet_name": r.excel_sheet_name,
        "excel_row_ref": r.excel_row_ref,
        "raw_payload_json": json.loads(r.raw_payload_json or "{}"),
        "normalized_payload_json": json.loads(r.normalized_payload_json or "{}"),
        "account_uid": r.account_uid,
        "abonent_id": r.abonent_id,
        "account_number": r.account_number,
        "payment_date": r.payment_date,
        "paid_date": r.paid_date,
        "payment_period": r.payment_period,
        "charge_year": r.charge_year,
        "charge_month": r.charge_month,
        "amount": r.amount,
        "source_index": r.source_index,
        "source_label": r.source_label,
        "fingerprint": r.fingerprint,
        "matched_payment_id": r.matched_payment_id,
        "applied_at": int(r.applied_at.timestamp() * 1000) if r.applied_at else 0,
        "status": r.status,
        "reason_code": r.reason_code,
        "reason_text": r.reason_text,
        "error_details_json": json.loads(r.error_details_json or "{}"),
    }


def _find_cell(header_map, row_values, *keys):
    for k in keys:
        if k in header_map:
            return row_values[header_map[k]] if header_map[k] < len(row_values) else None
    return None


def _parse_header_map(values):
    out = {}
    for i, v in enumerate(values):
        key = _norm_text(v).lower()
        if not key:
            continue
        normalized = re.sub(r"[^\wа-яА-ЯёЁ]+", " ", key, flags=re.UNICODE)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                out[canonical] = i
                break
        if normalized not in out:
            out[normalized] = i
    return out


def _parse_strict_header_map(values):
    out = {}
    for i, v in enumerate(values):
        key = _norm_text(v).lower()
        if not key:
            continue
        normalized = re.sub(r"[^\wа-яА-ЯёЁ]+", "_", key, flags=re.UNICODE)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        if normalized in IMPORT_REQUIRED_COLUMNS:
            out[normalized] = i
    return out


def _ensure_batch_transition(batch: ImportBatch, operation: str):
    allowed = IMPORT_ALLOWED_TRANSITIONS.get(operation, set())
    if batch.status not in allowed:
        return jsonify(
            ok=False,
            error="state_transition_forbidden",
            details={
                "operation": operation,
                "from_status": batch.status,
                "allowed_from": sorted(allowed),
            },
        ), 409
    return None


def _load_owner_sources(owner_id: str):
    row = KVStore.query.filter_by(owner=owner_id, k="payment_sources_v1").first()
    if not row or not row.v:
        return {}
    try:
        arr = json.loads(row.v)
    except Exception:
        return {}
    if not isinstance(arr, list):
        return {}
    out = {}
    for i, label in enumerate(arr, start=1):
        clean = _norm_text(label)
        if clean:
            out[i] = clean
    return out


def _extract_abonents_values(obj):
    if not isinstance(obj, dict):
        return []
    nested = obj.get("abonents")
    if isinstance(nested, dict):
        return nested.values()
    return obj.values()


def _extract_abonents_items(obj):
    if not isinstance(obj, dict):
        return []
    nested = obj.get("abonents")
    if isinstance(nested, dict):
        return list(nested.items())
    return list(obj.items())


def _find_owner_accounts(owner_id: str, account_uid: str, account_number: str):
    uid_norm = _norm_text(account_uid).lower()
    ls_norm = _norm_text(account_number)
    if not uid_norm:
        return {"matches": [], "uid_found": False}

    uid_hits = []
    matches = []
    for key in ("abonents_db_v1", "abonents_v1"):
        row = KVStore.query.filter_by(owner=owner_id, k=key).first()
        if not row or not row.v:
            continue
        try:
            obj = json.loads(row.v)
        except Exception:
            continue

        for ls_key, abonent in _extract_abonents_items(obj):
            if not isinstance(abonent, dict):
                continue
            candidate_uid = _norm_text(abonent.get("uid")).lower()
            if candidate_uid != uid_norm:
                continue
            uid_hits.append(abonent)
            key_ls = _norm_text(ls_key)
            fallback_ls = _norm_text(abonent.get("id"))
            if ls_norm and key_ls and key_ls == ls_norm:
                matches.append(abonent)
                continue
            if ls_norm and fallback_ls and fallback_ls == ls_norm:
                matches.append(abonent)
                continue
            if ls_norm:
                continue
            matches.append(abonent)

        if uid_hits:
            break

    return {"matches": matches, "uid_found": bool(uid_hits)}



def _summary_identity_value(abonent: dict, keys, fallback: str = ""):
    if not isinstance(abonent, dict):
        return fallback
    for key in keys:
        value = _norm_text(abonent.get(key))
        if value:
            return value
    return fallback


def _summary_identity_values(abonent: dict, keys):
    if not isinstance(abonent, dict):
        return {}
    return {key: _norm_text(abonent.get(key)) for key in keys if _norm_text(abonent.get(key))}




def _owner_abonents_db_v1_summary_targets(owner_id: str):
    targets = []
    seen_uids = set()

    row = KVStore.query.filter_by(owner=owner_id, k="abonents_db_v1").first()
    if not row or not row.v:
        return targets

    try:
        obj = json.loads(row.v)
    except Exception:
        return targets

    for ls_key, abonent in _extract_abonents_items(obj):
        if not isinstance(abonent, dict):
            continue
        account_uid = _norm_text(abonent.get("uid"))
        if not account_uid or account_uid in seen_uids:
            continue
        seen_uids.add(account_uid)

        key_account_number = _norm_text(ls_key)
        account_number = _summary_identity_value(
            abonent,
            ("account_number", "accountNumber", "ls", "id"),
            key_account_number,
        )
        abonent_id = _summary_identity_value(
            abonent,
            ("abonent_id", "abonentId", "id"),
            key_account_number or account_number,
        )

        targets.append({
            "abonent_id": abonent_id,
            "account_uid": account_uid,
            "account_number": account_number,
            "source": abonent,
            "identity": {
                "abonent_id": abonent_id,
                "account_uid": account_uid,
                "account_number": account_number,
                "fio": _summary_identity_value(abonent, ("fio", "fullName", "full_name")),
                "address": _summary_identity_value(abonent, ("address", "addr")),
                **_summary_identity_values(abonent, (
                    "calcStartDate", "calc_start_date",
                    "dateFrom", "date_from",
                    "responsibilityFrom", "respFrom",
                    "calcEndDate", "calc_end_date",
                    "dateTo", "date_to",
                    "responsibilityTo", "respTo",
                )),
            },
        })

    return targets


def _dirty_abonent_summary_payload(existing_summary: dict | None, reason: str):
    payload = existing_summary.copy() if isinstance(existing_summary, dict) else {}
    for key in (
        "totals",
        "total",
        "total_debt",
        "total_penalty",
        "total_accrued",
        "total_paid",
        "total_principal",
        "debt",
        "penalty",
        "principal",
    ):
        payload.pop(key, None)
    payload.update({
        "summary_status": "dirty",
        "summary_reason": _cache_reason("summary", "dirty", reason),
        "status": "dirty",
        "reason": _cache_reason("summary", "dirty", reason),
        "dirty_at": datetime.utcnow().isoformat() + "Z",
    })
    if "period" not in payload or not isinstance(payload.get("period"), dict):
        payload["period"] = {"from": None, "to": None}
    return payload

def _owner_abonent_summary_targets(owner_id: str):
    targets = []
    seen_uids = set()

    for key in ("abonents_db_v1", "abonents_v1"):
        row = KVStore.query.filter_by(owner=owner_id, k=key).first()
        if not row or not row.v:
            continue
        try:
            obj = json.loads(row.v)
        except Exception:
            continue

        for ls_key, abonent in _extract_abonents_items(obj):
            if not isinstance(abonent, dict):
                continue
            account_uid = _norm_text(abonent.get("uid"))
            if not account_uid or account_uid in seen_uids:
                continue
            seen_uids.add(account_uid)

            key_account_number = _norm_text(ls_key)
            account_number = _summary_identity_value(
                abonent,
                ("account_number", "accountNumber", "ls", "id"),
                key_account_number,
            )
            abonent_id = _summary_identity_value(
                abonent,
                ("abonent_id", "abonentId", "id"),
                key_account_number or account_number,
            )

            targets.append({
                "abonent_id": abonent_id,
                "account_uid": account_uid,
                "account_number": account_number,
                "source": abonent,
                "identity": {
                    "abonent_id": abonent_id,
                    "account_uid": account_uid,
                    "account_number": account_number,
                    "fio": _summary_identity_value(abonent, ("fio", "fullName", "full_name")),
                    "address": _summary_identity_value(abonent, ("address", "addr")),
                    **_summary_identity_values(abonent, (
                        "calcStartDate", "calc_start_date",
                        "dateFrom", "date_from",
                        "responsibilityFrom", "respFrom",
                        "calcEndDate", "calc_end_date",
                        "dateTo", "date_to",
                        "responsibilityTo", "respTo",
                    )),
                },
            })

        if targets:
            break

    return targets


def _build_missing_abonent_summary(target: dict):
    identity = target.get("identity") if isinstance(target.get("identity"), dict) else {}
    return {
        "status": "missing",
        "reason": "SUMMARY_NOT_BUILT",
        "summary_status": "missing",
        "summary_reason": "SUMMARY_NOT_BUILT",
        "abonent": {
            "abonent_id": _norm_text(identity.get("abonent_id") or target.get("abonent_id")),
            "account_uid": _norm_text(identity.get("account_uid") or target.get("account_uid")),
            "account_number": _norm_text(identity.get("account_number") or target.get("account_number")),
            "fio": _norm_text(identity.get("fio")),
            "address": _norm_text(identity.get("address")),
        },
        "period": {
            "from": None,
            "to": None,
        },
    }


def _build_batch_recalc_error_summary(target: dict, existing_summary: dict | None, reason: str):
    identity = target.get("identity") if isinstance(target.get("identity"), dict) else {}
    existing = existing_summary if isinstance(existing_summary, dict) else {}
    period = existing.get("period") if isinstance(existing.get("period"), dict) else {}
    clean_reason = _norm_text(reason) or "BATCH_RECALC_FAILED"
    return {
        "status": "error",
        "reason": clean_reason,
        "summary_status": "error",
        "summary_reason": clean_reason,
        "abonent": {
            "abonent_id": _norm_text(identity.get("abonent_id") or target.get("abonent_id") or existing.get("abonent_id")),
            "account_uid": _norm_text(identity.get("account_uid") or target.get("account_uid") or existing.get("account_uid") or existing.get("uid")),
            "account_number": _norm_text(identity.get("account_number") or target.get("account_number") or existing.get("account_number")),
            "fio": _norm_text(identity.get("fio") or existing.get("fio")),
            "address": _norm_text(identity.get("address") or existing.get("address")),
        },
        "account_uid": _norm_text(target.get("account_uid") or existing.get("account_uid") or existing.get("uid")),
        "uid": _norm_text(target.get("account_uid") or existing.get("account_uid") or existing.get("uid")),
        "account_number": _norm_text(target.get("account_number") or existing.get("account_number")),
        "abonent_id": _norm_text(target.get("abonent_id") or existing.get("abonent_id") or existing.get("id")),
        "period": {
            "from": _norm_text(period.get("from")),
            "to": _norm_text(period.get("to")),
        },
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def _upsert_abonent_summary_payload(owner_id: str, account_uid: str, target: dict, summary: dict):
    account_uid = _norm_text(account_uid)
    if not account_uid:
        return None
    abonent_id = _norm_text(target.get("abonent_id") if isinstance(target, dict) else "")
    account_number = _norm_text(target.get("account_number") if isinstance(target, dict) else "")
    row = AbonentSummary.query.filter_by(owner_id=owner_id, account_uid=account_uid).order_by(AbonentSummary.id.asc()).first()
    if row:
        row.abonent_id = abonent_id or row.abonent_id
        row.account_number = account_number or row.account_number
        _set_abonent_summary_json(row, target, summary)
    else:
        row = AbonentSummary(
            owner_id=owner_id,
            abonent_id=abonent_id,
            account_uid=account_uid,
            account_number=account_number,
        )
        _set_abonent_summary_json(row, target, summary)
        db.session.add(row)
    return row


def _summary_status_from_payload(summary: dict | None):
    if not isinstance(summary, dict):
        return "missing"
    scope = _norm_text(summary.get("summary_scope") or summary.get("report_scope")).lower()
    if scope in {"period", "report"}:
        return "missing"
    status = _norm_text(summary.get("summary_status") or summary.get("status")).lower()
    reason = _norm_text(summary.get("summary_reason") or summary.get("reason"))
    if status == "dirty" and reason == "CALC_PERIOD_CHANGED":
        return "missing"
    if status in {"fresh", "dirty", "missing", "error", "invalid"}:
        return status
    return "missing"


def _summary_finite_number(value):
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    try:
        n = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return False
    return n.is_finite()


def _pick_summary_total(summary: dict, totals: dict, keys: tuple[str, ...]):
    for key in keys:
        if key.startswith("totals."):
            value = totals.get(key.split(".", 1)[1])
        else:
            value = summary.get(key)
        if value is not None and not (isinstance(value, str) and not value.strip()):
            return value
    return None


def _fresh_totals_validation_reason(summary: dict | None):
    if _summary_status_from_payload(summary) != "fresh":
        return ""
    if not isinstance(summary, dict):
        return "FRESH_TOTALS_MISSING"

    totals = summary.get("totals")
    if not isinstance(totals, dict):
        return "FRESH_TOTALS_MISSING"

    required = {
        "debt": ("totals.debt", "totals.total", "totals.total_debt"),
        "penalty": ("totals.penalty", "totals.total_penalty"),
        "accrued": ("totals.accrued", "totals.total_accrued"),
        "paid": ("totals.paid", "totals.total_paid"),
    }
    missing = False
    invalid = False
    for keys in required.values():
        value = _pick_summary_total(summary, totals, keys)
        if value is None or (isinstance(value, str) and not value.strip()):
            missing = True
            continue
        if not _summary_finite_number(value):
            invalid = True

    if missing:
        return "FRESH_TOTALS_MISSING"
    if invalid:
        return "FRESH_TOTALS_INVALID"
    return ""


def _summary_with_validated_fresh_totals(summary: dict | None):
    payload = summary.copy() if isinstance(summary, dict) else {}
    reason = _fresh_totals_validation_reason(payload)
    if reason:
        payload["summary_status"] = "error"
        payload["summary_reason"] = reason
        payload["status"] = "error"
        payload["reason"] = reason
    return payload


def _summary_date_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text_value = _norm_text(value)
    if not text_value:
        return ""
    match = re.search(r"\d{4}-\d{2}-\d{2}", text_value)
    if match:
        return match.group(0)
    return ""


def _pick_summary_boundary(target: dict | None, keys: tuple[str, ...]):
    if not isinstance(target, dict):
        return ""
    identity = target.get("identity") if isinstance(target.get("identity"), dict) else {}
    source = target.get("source") if isinstance(target.get("source"), dict) else {}
    for container in (target, identity, source):
        for key in keys:
            value = _summary_date_iso(container.get(key))
            if value:
                return value
    return ""


def _is_legacy_period_summary_contamination(summary: dict | None, target: dict | None):
    if not isinstance(summary, dict):
        return False
    status = _norm_text(summary.get("summary_status") or summary.get("status")).lower()
    if status != "fresh":
        return False
    scope = _norm_text(summary.get("summary_scope") or summary.get("report_scope")).lower()
    if scope in {"period", "report"}:
        return False
    period = summary.get("period")
    if not isinstance(period, dict):
        return False
    summary_from = _summary_date_iso(period.get("from"))
    summary_to = _summary_date_iso(period.get("to"))
    if not summary_from or not summary_to:
        return False

    canonical_from = _pick_summary_boundary(target, (
        "calcStartDate", "calc_start_date",
        "dateFrom", "date_from",
        "responsibilityFrom", "respFrom",
    ))
    canonical_to = _pick_summary_boundary(target, (
        "calcEndDate", "calc_end_date",
        "dateTo", "date_to",
        "responsibilityTo", "respTo",
    ))
    if not canonical_from and not canonical_to:
        try:
            app.logger.info(
                "[summary][legacy-period-contamination][skip] owner_id=%s account_uid=%s abonent_id=%s summary_from=%s summary_to=%s reason=CANONICAL_BOUNDARIES_UNKNOWN",
                _norm_text(target.get("owner_id")) if isinstance(target, dict) else "",
                _norm_text(target.get("account_uid")) if isinstance(target, dict) else "",
                _norm_text(target.get("abonent_id")) if isinstance(target, dict) else "",
                summary_from,
                summary_to,
            )
        except Exception:
            pass
        return False

    return bool(
        (canonical_from and summary_from != canonical_from) or
        (canonical_to and summary_to != canonical_to)
    )


def _log_legacy_period_summary_contamination(owner_id: str, account_uid: str, abonent_id: str, summary: dict, target: dict | None):
    period = summary.get("period") if isinstance(summary.get("period"), dict) else {}
    try:
        app.logger.warning(
            "[summary][legacy-period-contamination] owner_id=%s account_uid=%s abonent_id=%s summary_from=%s summary_to=%s canonical_from=%s canonical_to=%s",
            _norm_text(owner_id),
            _norm_text(account_uid),
            _norm_text(abonent_id),
            _summary_date_iso(period.get("from")),
            _summary_date_iso(period.get("to")),
            _pick_summary_boundary(target, (
                "calcStartDate", "calc_start_date",
                "dateFrom", "date_from",
                "responsibilityFrom", "respFrom",
            )),
            _pick_summary_boundary(target, (
                "calcEndDate", "calc_end_date",
                "dateTo", "date_to",
                "responsibilityTo", "respTo",
            )),
        )
    except Exception:
        pass


def _summary_without_stale_totals(summary: dict | None, target: dict | None = None, owner_id: str = ""):
    payload = _summary_with_validated_fresh_totals(summary)
    if _is_legacy_period_summary_contamination(payload, target):
        _log_legacy_period_summary_contamination(
            owner_id,
            _norm_text(target.get("account_uid")) if isinstance(target, dict) else "",
            _norm_text(target.get("abonent_id")) if isinstance(target, dict) else "",
            payload,
            target,
        )
        payload["summary_status"] = "missing"
        payload["status"] = "missing"
        payload["summary_reason"] = "PERIOD_SUMMARY_LEGACY"
        payload["reason"] = "PERIOD_SUMMARY_LEGACY"
    status = _summary_status_from_payload(payload)
    scope = _norm_text(payload.get("summary_scope") or payload.get("report_scope")).lower()
    if scope in {"period", "report"}:
        payload["summary_reason"] = "PERIOD_SUMMARY_IGNORED"
        payload["reason"] = "PERIOD_SUMMARY_IGNORED"
    payload["summary_status"] = status
    payload["status"] = status
    if status != "fresh":
        for key in (
            "totals",
            "total",
            "total_debt",
            "total_penalty",
            "total_accrued",
            "total_paid",
            "total_principal",
            "debt",
            "penalty",
            "principal",
            "nachisleno",
            "oplacheno",
            "accrued",
            "paid",
            "accrued_total",
            "paid_total",
            "accruedTotal",
            "paidTotal",
        ):
            payload.pop(key, None)
    return payload


def _summary_from_row_or_missing(row: AbonentSummary | None, target: dict):
    if not row:
        return _build_missing_abonent_summary(target)
    try:
        summary = json.loads(row.summary_json or "{}")
    except (TypeError, ValueError):
        summary = {"summary_status": "error", "summary_reason": "SUMMARY_JSON_INVALID"}
    return _summary_without_stale_totals(summary, target, row.owner_id)


def _batch_job_payload(job: RecalcBatchJob):
    status = _norm_text(job.status) or "queued"
    done = int(job.processed_count or 0)
    total = int(job.total_count or 0)
    errors = int(job.error_count or 0) + int(job.skipped_count or 0)
    return {
        "id": int(job.id),
        "status": status,
        "reason": _norm_text(job.reason),
        "total_count": total,
        "processed_count": done,
        "fresh_count": int(job.fresh_count or 0),
        "error_count": int(job.error_count or 0),
        "skipped_count": int(job.skipped_count or 0),
        "started_at": job.started_at.isoformat() + "Z" if job.started_at else None,
        "finished_at": job.finished_at.isoformat() + "Z" if job.finished_at else None,
        "total": total,
        "done": done,
        "errors": errors,
    }


def _batch_job_status_response(job: RecalcBatchJob):
    payload = _batch_job_payload(job)
    items = RecalcBatchJobItem.query.filter_by(job_id=job.id, owner_id=job.owner_id).order_by(RecalcBatchJobItem.id.asc()).all()
    affected_uids = [_norm_text(x.account_uid) for x in items if _norm_text(x.account_uid)]
    errors_by_uid = {
        _norm_text(x.account_uid): _norm_text(x.error_message) or _norm_text(x.summary_reason) or "UID_PROCESSING_FAILED"
        for x in items
        if _norm_text(x.account_uid) and (_norm_text(x.status) == "error" or _norm_text(x.error_message))
    }
    sample_errors = [
        {"account_uid": uid, "error": error}
        for uid, error in list(errors_by_uid.items())[:10]
    ]
    return {
        "ok": True,
        "job": payload,
        "job_id": payload["id"],
        "status": payload["status"],
        "total": payload["total_count"],
        "processed": payload["processed_count"],
        "fresh": payload["fresh_count"],
        "error": payload["error_count"],
        "skipped": payload["skipped_count"],
        "started_at": payload["started_at"],
        "finished_at": payload["finished_at"],
        "message": _norm_text(job.error_message) or "",
        "affected_uids": affected_uids,
        "errors_by_uid": errors_by_uid,
        "sample_errors": sample_errors,
        "items": [{
            "account_uid": _norm_text(x.account_uid),
            "status": _norm_text(x.status),
            "summary_status": _norm_text(x.summary_status),
            "summary_reason": _norm_text(x.summary_reason),
            "error_message": _norm_text(x.error_message),
        } for x in items],
    }


def _recalc_batch_process_job(owner_id: str, job: RecalcBatchJob, step_limit: int = 25):
    if not job or job.owner_id != owner_id:
        return
    if job.status not in {"queued", "running"}:
        return
    if job.status == "queued":
        job.status = "running"
        if not job.started_at:
            job.started_at = datetime.utcnow()
        db.session.commit()
    items = (
        RecalcBatchJobItem.query
        .filter_by(job_id=job.id, owner_id=owner_id)
        .filter(RecalcBatchJobItem.status.in_(["queued", "running"]))
        .order_by(RecalcBatchJobItem.id.asc())
        .limit(max(1, int(step_limit)))
        .all()
    )
    if not items:
        result_count = int(job.fresh_count or 0) + int(job.error_count or 0) + int(job.skipped_count or 0)
        if int(job.processed_count or 0) >= int(job.total_count or 0) and result_count >= int(job.total_count or 0):
            job.status = "completed"
            if not job.finished_at:
                job.finished_at = datetime.utcnow()
            db.session.commit()
        elif int(job.processed_count or 0) >= int(job.total_count or 0):
            job.status = "failed"
            job.error_message = "BATCH_COUNTERS_INCONSISTENT"
            if not job.finished_at:
                job.finished_at = datetime.utcnow()
            db.session.commit()
        return
    targets = _owner_abonent_summary_targets(owner_id)
    targets_by_uid = {_norm_text(t.get("account_uid")): t for t in targets if _norm_text(t.get("account_uid"))}
    for item in items:
        item.status = "running"
        if not item.started_at:
            item.started_at = datetime.utcnow()
        db.session.commit()
        try:
            row = AbonentSummary.query.filter_by(owner_id=owner_id, account_uid=item.account_uid).order_by(AbonentSummary.id.asc()).first()
            target = targets_by_uid.get(_norm_text(item.account_uid)) or {"account_uid": item.account_uid, "abonent_id": "", "account_number": "", "identity": {}}
            summary = _summary_from_row_or_missing(row, target)
            summary = _summary_with_validated_fresh_totals(summary)
            s_status = _summary_status_from_payload(summary)
            s_reason = _norm_text(summary.get("summary_reason") or summary.get("reason") or "")
            if s_status == "fresh":
                item.status = "fresh"
                job.fresh_count = int(job.fresh_count or 0) + 1
            else:
                result_reason = s_reason or s_status or "SUMMARY_NOT_FRESH"
                if s_status in {"dirty", "missing"}:
                    result_reason = "BATCH_RECALC_NOT_AVAILABLE" + (":" + result_reason if result_reason else "")
                error_summary = _build_batch_recalc_error_summary(target, summary, result_reason)
                _upsert_abonent_summary_payload(owner_id, item.account_uid, target, error_summary)
                item.status = "error"
                item.error_message = result_reason
                s_status = "error"
                s_reason = result_reason
                job.error_count = int(job.error_count or 0) + 1
            item.summary_status = s_status
            item.summary_reason = s_reason
        except Exception as exc:
            item.status = "error"
            item.error_message = f"UID_PROCESSING_FAILED: {exc}"
            item.summary_status = "error"
            item.summary_reason = "UID_PROCESSING_FAILED"
            job.error_count = int(job.error_count or 0) + 1
        item.finished_at = datetime.utcnow()
        job.processed_count = int(job.processed_count or 0) + 1
        db.session.commit()
    remaining = RecalcBatchJobItem.query.filter_by(job_id=job.id, owner_id=owner_id).filter(RecalcBatchJobItem.status.in_(["queued", "running"])).count()
    result_count = int(job.fresh_count or 0) + int(job.error_count or 0) + int(job.skipped_count or 0)
    if remaining == 0 and int(job.processed_count or 0) >= int(job.total_count or 0) and result_count >= int(job.total_count or 0):
        job.status = "completed"
        job.finished_at = datetime.utcnow()
        db.session.commit()
    elif remaining == 0 and int(job.processed_count or 0) >= int(job.total_count or 0):
        job.status = "failed"
        job.error_message = "BATCH_COUNTERS_INCONSISTENT"
        job.finished_at = datetime.utcnow()
        db.session.commit()


def _client_recalc_job_progress_response(job: RecalcBatchJob):
    return jsonify(**_batch_job_status_response(job))


def _summary_uid_from_payload(summary: dict):
    if not isinstance(summary, dict):
        return ""
    abonent = summary.get("abonent") if isinstance(summary.get("abonent"), dict) else {}
    return _norm_text(
        summary.get("account_uid")
        or summary.get("uid")
        or summary.get("accountUid")
        or abonent.get("account_uid")
        or abonent.get("uid")
        or abonent.get("accountUid")
    )


def _client_recalc_summary_payload(status: str, summary: dict | None, target: dict, account_uid: str, reason: str):
    payload = dict(summary) if isinstance(summary, dict) else {}
    clean_status = _norm_text(status).lower()
    if clean_status not in {"fresh", "error", "skipped"}:
        clean_status = "error"
    clean_reason = _norm_text(reason)
    if clean_status == "fresh":
        clean_reason = clean_reason or _norm_text(payload.get("summary_reason") or payload.get("reason")) or "RECALC_OK"
    elif clean_status == "skipped":
        clean_reason = clean_reason or "SKIPPED_BY_CLIENT"
    else:
        clean_reason = clean_reason or _norm_text(payload.get("summary_reason") or payload.get("reason")) or "CLIENT_RECALC_FAILED"

    payload["summary_status"] = clean_status
    payload["status"] = clean_status
    payload["summary_reason"] = clean_reason
    payload["reason"] = clean_reason
    payload["calculation_source"] = "CLIENT_CALCULATED_SUMMARY"
    payload["account_uid"] = account_uid
    payload["uid"] = account_uid
    payload["account_number"] = _norm_text(target.get("account_number")) or _norm_text(payload.get("account_number"))
    payload["abonent_id"] = _norm_text(target.get("abonent_id")) or _norm_text(payload.get("abonent_id"))

    if clean_status != "fresh":
        for key in (
            "totals",
            "total",
            "total_debt",
            "total_penalty",
            "total_accrued",
            "total_paid",
            "penalty_debt",
            "main_debt",
            "penalty",
            "debt",
            "accrued",
            "paid",
        ):
            payload.pop(key, None)
    return payload, clean_status, clean_reason


@app.get("/api/recalc_batch_job/<int:job_id>/next_uid")
def client_recalc_batch_job_next_uid(job_id: int):
    user, err = _require_user()
    if err:
        return err
    job = RecalcBatchJob.query.filter_by(id=job_id, owner_id=user.id).first()
    if not job:
        return jsonify(ok=False, error="job_not_found"), 404
    if job.status not in {"queued", "running"}:
        return jsonify(ok=False, error="job_not_active", status=_norm_text(job.status)), 400

    if job.status == "queued":
        job.status = "running"
        if not job.started_at:
            job.started_at = datetime.utcnow()

    item = (
        RecalcBatchJobItem.query
        .filter_by(job_id=job.id, owner_id=user.id, status="queued")
        .order_by(RecalcBatchJobItem.id.asc())
        .first()
    )
    if item:
        item.status = "running"
        item.started_at = datetime.utcnow()
        db.session.commit()
        targets_by_uid = _owner_recalc_targets_by_uid(user.id)
        target = targets_by_uid.get(_norm_text(item.account_uid)) or {}
        return jsonify(
            ok=True,
            job_id=int(job.id),
            item_id=int(item.id),
            account_uid=_norm_text(item.account_uid),
            account_number=_norm_text(target.get("account_number")),
        )

    running = RecalcBatchJobItem.query.filter_by(job_id=job.id, owner_id=user.id, status="running").count()
    if running:
        db.session.commit()
        return jsonify(ok=True, status="retry", reason="no_available_uid", job_id=int(job.id))

    job.status = "done"
    if not job.finished_at:
        job.finished_at = datetime.utcnow()
    db.session.commit()
    return jsonify(ok=True, status="done", reason="all_done", job_id=int(job.id), all_done=True)


@app.post("/api/recalc_batch_job/<int:job_id>/complete_uid")
def client_recalc_batch_job_complete_uid(job_id: int):
    user, err = _require_user()
    if err:
        return err
    job = RecalcBatchJob.query.filter_by(id=job_id, owner_id=user.id).first()
    if not job:
        return jsonify(ok=False, error="job_not_found"), 404

    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify(ok=False, error="invalid_payload"), 400
    try:
        item_id = int(body.get("item_id"))
    except (TypeError, ValueError):
        return jsonify(ok=False, error="item_id_required"), 400

    item = RecalcBatchJobItem.query.filter_by(id=item_id, job_id=job.id, owner_id=user.id).first()
    if not item:
        return jsonify(ok=False, error="item_not_found"), 404
    if item.status != "running":
        return jsonify(ok=False, error="item_not_running", item_status=_norm_text(item.status)), 409

    status = _norm_text(body.get("status")).lower()
    if status not in {"fresh", "error", "skipped"}:
        return jsonify(ok=False, error="invalid_status"), 400

    summary = body.get("summary") if isinstance(body.get("summary"), dict) else {}
    summary_uid = _summary_uid_from_payload(summary)
    if summary_uid and summary_uid != _norm_text(item.account_uid):
        return jsonify(ok=False, error="uid_mismatch"), 400

    targets_by_uid = _owner_recalc_targets_by_uid(user.id)
    target = targets_by_uid.get(_norm_text(item.account_uid)) or {
        "owner_id": user.id,
        "account_uid": item.account_uid,
        "account_number": "",
        "abonent_id": "",
        "identity": {"account_uid": item.account_uid},
    }
    error_reason = _norm_text(body.get("error_reason"))
    payload, summary_status, summary_reason = _client_recalc_summary_payload(
        status,
        summary,
        target,
        _norm_text(item.account_uid),
        error_reason,
    )

    if summary_status in {"fresh", "error"}:
        row = AbonentSummary.query.filter_by(owner_id=user.id, account_uid=item.account_uid).order_by(AbonentSummary.id.asc()).first()
        if row:
            row.abonent_id = _norm_text(target.get("abonent_id")) or row.abonent_id
            row.account_number = _norm_text(target.get("account_number")) or row.account_number
        else:
            row = AbonentSummary(
                owner_id=user.id,
                abonent_id=_norm_text(target.get("abonent_id")),
                account_uid=_norm_text(item.account_uid),
                account_number=_norm_text(target.get("account_number")),
            )
            db.session.add(row)
        _set_abonent_summary_json(row, target, payload)

    item.status = summary_status
    item.summary_status = summary_status
    item.summary_reason = summary_reason
    item.error_message = "" if summary_status == "fresh" else summary_reason
    item.finished_at = datetime.utcnow()
    job.processed_count = int(job.processed_count or 0) + 1
    if summary_status == "fresh":
        job.fresh_count = int(job.fresh_count or 0) + 1
    elif summary_status == "skipped":
        job.skipped_count = int(job.skipped_count or 0) + 1
    else:
        job.error_count = int(job.error_count or 0) + 1

    remaining = RecalcBatchJobItem.query.filter_by(job_id=job.id, owner_id=user.id).filter(RecalcBatchJobItem.status.in_(["queued", "running"])).count()
    if remaining == 0:
        job.status = "done"
        job.finished_at = datetime.utcnow()
    elif job.status == "queued":
        job.status = "running"
        if not job.started_at:
            job.started_at = datetime.utcnow()

    db.session.commit()
    db.session.refresh(job)
    return _client_recalc_job_progress_response(job)


def _recalc_normalize_uids(raw_uids):
    out = []
    seen = set()
    for value in (raw_uids or []):
        uid = _norm_text(value)
        if not uid or uid in seen:
            continue
        seen.add(uid)
        out.append(uid)
    out.sort()
    return out


def _recalc_job_fingerprint(owner_id: str, reason: str, normalized_uids):
    raw = "|".join([_norm_text(owner_id), _norm_text(reason) or "MANUAL_RECALC", ",".join(normalized_uids)])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _recalc_recover_stale_running_jobs(owner_id: str):
    now = datetime.utcnow()
    jobs = RecalcBatchJob.query.filter_by(owner_id=owner_id).filter(RecalcBatchJob.status.in_(["queued", "running"])).all()
    for job in jobs:
        anchor = job.started_at or job.created_at
        if not anchor:
            continue
        age = (now - anchor).total_seconds()
        if age <= RECALC_BATCH_RUNNING_TTL_SECONDS:
            continue
        job.status = "stale"
        job.error_message = "STALE_RUNNING_RECOVERED"
        job.finished_at = now
    db.session.flush()


def _recalc_cleanup_old_jobs(owner_id: str):
    now = datetime.utcnow()
    threshold = now.timestamp() - (RECALC_BATCH_RETENTION_DAYS * 86400)
    final_jobs = RecalcBatchJob.query.filter_by(owner_id=owner_id).filter(RecalcBatchJob.status.in_(["completed", "failed", "stale"])).order_by(RecalcBatchJob.id.desc()).all()
    keep_ids = set(j.id for j in final_jobs[:RECALC_BATCH_KEEP_PER_OWNER])
    for job in final_jobs[RECALC_BATCH_KEEP_PER_OWNER:]:
        finished_ts = (job.finished_at or job.created_at)
        should_delete = True
        if finished_ts:
            should_delete = finished_ts.timestamp() < threshold
        if not should_delete:
            continue
        db.session.query(RecalcBatchJobItem).filter_by(job_id=job.id, owner_id=owner_id).delete()
        db.session.delete(job)


def _owner_recalc_targets_by_uid(owner_id: str):
    targets = {}
    for target in _owner_abonent_summary_targets(owner_id):
        uid = _norm_text(target.get("account_uid"))
        if uid:
            targets[uid] = target

    cfg = _abonents_api_select_config()
    if cfg is None:
        return targets
    uid_col = cfg.get("uid_col")
    owner_col = cfg.get("owner_col")
    if not uid_col or not owner_col:
        return targets

    select_parts = [f"a.{_sql_ident(uid_col)} AS account_uid"]
    account_col = cfg.get("account_col")
    abonent_id_col = cfg.get("abonent_id_col")
    fio_col = cfg.get("fio_col")
    if account_col:
        select_parts.append(f"a.{_sql_ident(account_col)} AS account_number")
    else:
        select_parts.append("'' AS account_number")
    if abonent_id_col:
        select_parts.append(f"a.{_sql_ident(abonent_id_col)} AS abonent_id")
    else:
        select_parts.append("'' AS abonent_id")
    if fio_col:
        select_parts.append(f"a.{_sql_ident(fio_col)} AS fio")
    else:
        select_parts.append("'' AS fio")

    rows = db.session.execute(
        text(
            f"""
            SELECT {", ".join(select_parts)}
            FROM {_sql_ident(cfg["abonent_table"])} a
            WHERE a.{_sql_ident(owner_col)} = :owner
            """
        ),
        {"owner": owner_id},
    ).all()
    for row in rows:
        data = dict(row._mapping)
        uid = _norm_text(data.get("account_uid"))
        if not uid or uid in targets:
            continue
        targets[uid] = {
            "owner_id": owner_id,
            "account_uid": uid,
            "account_number": _norm_text(data.get("account_number")),
            "abonent_id": _norm_text(data.get("abonent_id")),
            "identity": {
                "account_uid": uid,
                "account_number": _norm_text(data.get("account_number")),
                "abonent_id": _norm_text(data.get("abonent_id")),
                "fio": _norm_text(data.get("fio")),
            },
        }
    return targets


def _recalc_batch_create_job(owner_id: str, user_id: str, raw_uids, reason: str):
    requested = _recalc_normalize_uids(raw_uids)
    if not requested:
        return None, {"requested": 0, "accepted": 0, "skipped": 0}
    if len(requested) > RECALC_BATCH_MAX_UIDS:
        return "TOO_MANY_UIDS", {"max_uids": RECALC_BATCH_MAX_UIDS, "requested": len(requested)}
    reason_norm = _norm_text(reason) or "MANUAL_RECALC"

    _recalc_recover_stale_running_jobs(owner_id)

    fp = _recalc_job_fingerprint(owner_id, reason_norm, requested)
    active_jobs = RecalcBatchJob.query.filter_by(owner_id=owner_id, reason=reason_norm).filter(RecalcBatchJob.status.in_(["queued", "running"])).order_by(RecalcBatchJob.id.desc()).all()
    for active_job in active_jobs:
        item_uids = [
            _norm_text(x.account_uid)
            for x in RecalcBatchJobItem.query.filter_by(job_id=active_job.id, owner_id=owner_id).all()
            if _norm_text(x.account_uid)
        ]
        if _recalc_job_fingerprint(owner_id, reason_norm, sorted(set(item_uids))) == fp:
            db.session.commit()
            return active_job, {"requested": len(requested), "accepted": int(active_job.total_count or 0), "skipped": int(active_job.skipped_count or 0)}

    targets_by_uid = _owner_recalc_targets_by_uid(owner_id)
    active_uid_rows = (
        db.session.query(RecalcBatchJobItem.account_uid)
        .join(RecalcBatchJob, RecalcBatchJob.id == RecalcBatchJobItem.job_id)
        .filter(RecalcBatchJob.owner_id == owner_id)
        .filter(RecalcBatchJob.status.in_(["queued", "running"]))
        .filter(RecalcBatchJobItem.status.in_(["queued", "running"]))
        .all()
    )
    active_uids = {_norm_text(row[0]) for row in active_uid_rows if _norm_text(row[0])}
    accepted = [uid for uid in requested if uid in targets_by_uid and uid not in active_uids]
    skipped = len(requested) - len(accepted)
    job = RecalcBatchJob(owner_id=owner_id, requested_by=user_id, reason=reason_norm, status="queued", total_count=len(accepted), skipped_count=skipped)
    db.session.add(job)
    db.session.flush()
    for uid in accepted:
        db.session.add(RecalcBatchJobItem(job_id=job.id, owner_id=owner_id, account_uid=uid, status="queued"))

    _recalc_cleanup_old_jobs(owner_id)
    db.session.commit()
    return job, {"requested": len(requested), "accepted": len(accepted), "skipped": skipped}


def _bulk_verify_json_load(raw_value, default=None):
    if default is None:
        default = {}
    try:
        parsed = json.loads(raw_value or "{}")
    except (TypeError, ValueError):
        return default
    return parsed if isinstance(parsed, dict) else default


def _bulk_verify_pick(summary: dict | None, keys: tuple[str, ...]):
    summary = summary if isinstance(summary, dict) else {}
    totals = summary.get("totals") if isinstance(summary.get("totals"), dict) else {}
    period = summary.get("period") if isinstance(summary.get("period"), dict) else {}
    for key in keys:
        if key.startswith("totals."):
            value = totals.get(key.split(".", 1)[1])
        elif key.startswith("period."):
            value = period.get(key.split(".", 1)[1])
        else:
            value = summary.get(key)
        if value is not None and not (isinstance(value, str) and not value.strip()):
            return value
    return None


def _bulk_verify_decimal_text(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        num = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return str(value)
    if not num.is_finite():
        return str(value)
    return str(num.quantize(Decimal("0.01")))


def _bulk_verify_text(value):
    if value is None:
        return None
    return _norm_text(value)


def _bulk_verify_summary_view(summary: dict | None):
    summary = summary if isinstance(summary, dict) else {}
    return {
        "total_accrued": _bulk_verify_decimal_text(_bulk_verify_pick(summary, ("total_accrued", "totals.total_accrued", "totals.accrued"))),
        "total_paid": _bulk_verify_decimal_text(_bulk_verify_pick(summary, ("total_paid", "totals.total_paid", "totals.paid"))),
        "main_debt": _bulk_verify_decimal_text(_bulk_verify_pick(summary, ("main_debt", "principal", "total_principal", "totals.principal"))),
        "penalty_debt": _bulk_verify_decimal_text(_bulk_verify_pick(summary, ("penalty_debt", "penalty", "total_penalty", "totals.penalty", "totals.total_penalty"))),
        "total_debt": _bulk_verify_decimal_text(_bulk_verify_pick(summary, ("total_debt", "total", "totals.total_debt", "totals.total", "totals.debt"))),
        "period_from": _bulk_verify_text(_bulk_verify_pick(summary, ("period_from", "period_start", "start_date", "period.from"))),
        "period_to": _bulk_verify_text(_bulk_verify_pick(summary, ("period_to", "period_end", "end_date", "period.to"))),
        "input_hash": _bulk_verify_text(_bulk_verify_pick(summary, ("input_hash",))),
        "version": _bulk_verify_text(_bulk_verify_pick(summary, ("version", "calc_engine_version", "engine_version"))),
    }


def _bulk_verify_summary_from_snapshot(snapshot: dict | None):
    src = snapshot if isinstance(snapshot, dict) else {}
    totals = src.get("totals") if isinstance(src.get("totals"), dict) else {}
    return {
        "summary_status": _norm_text(src.get("summary_status") or src.get("status") or "fresh").lower(),
        "summary_reason": _norm_text(src.get("summary_reason") or src.get("reason") or "OK"),
        "totals": totals,
        "total_accrued": _bulk_verify_pick(src, ("total_accrued", "totals.total_accrued", "totals.accrued")),
        "total_paid": _bulk_verify_pick(src, ("total_paid", "totals.total_paid", "totals.paid")),
        "main_debt": _bulk_verify_pick(src, ("main_debt", "principal", "total_principal", "totals.principal")),
        "penalty_debt": _bulk_verify_pick(src, ("penalty_debt", "penalty", "total_penalty", "totals.penalty", "totals.total_penalty")),
        "total_debt": _bulk_verify_pick(src, ("total_debt", "total", "totals.total_debt", "totals.total", "totals.debt")),
        "period": src.get("period") if isinstance(src.get("period"), dict) else {},
        "input_hash": _norm_text(src.get("input_hash")),
        "version": _norm_text(src.get("version") or src.get("calc_engine_version") or src.get("engine_version")),
    }


def _bulk_verify_diff(old_summary: dict | None, new_summary: dict | None):
    old_view = _bulk_verify_summary_view(old_summary)
    new_view = _bulk_verify_summary_view(new_summary)
    diff = {}
    for key in ("total_accrued", "total_paid", "main_debt", "penalty_debt", "total_debt", "period_from", "period_to", "input_hash", "version"):
        if old_view.get(key) != new_view.get(key):
            diff[key] = {"old": old_view.get(key), "new": new_view.get(key)}
    return old_view, new_view, diff


def _bulk_verify_active_uids(owner_id: str):
    active = set()
    recalc_rows = (
        db.session.query(RecalcBatchJobItem.account_uid)
        .join(RecalcBatchJob, RecalcBatchJob.id == RecalcBatchJobItem.job_id)
        .filter(RecalcBatchJob.owner_id == owner_id)
        .filter(RecalcBatchJob.status.in_(["queued", "running"]))
        .filter(RecalcBatchJobItem.status.in_(["queued", "running"]))
        .all()
    )
    active.update(_norm_text(row[0]) for row in recalc_rows if _norm_text(row[0]))
    verify_rows = (
        db.session.query(BulkCalcVerifyJobItem.account_uid)
        .join(BulkCalcVerifyJob, BulkCalcVerifyJob.id == BulkCalcVerifyJobItem.job_id)
        .filter(BulkCalcVerifyJob.owner_id == owner_id)
        .filter(BulkCalcVerifyJob.status.in_(["queued", "running"]))
        .filter(BulkCalcVerifyJobItem.status.in_(["queued", "running"]))
        .all()
    )
    active.update(_norm_text(row[0]) for row in verify_rows if _norm_text(row[0]))
    lock_rows = RecalcUidLock.query.filter_by(owner_id=owner_id, status="running").all()
    now = datetime.utcnow()
    for row in lock_rows:
        age = (now - (row.started_at or now)).total_seconds()
        if age <= RECALC_BATCH_RUNNING_TTL_SECONDS and _norm_text(row.abonent_uid):
            active.add(_norm_text(row.abonent_uid))
    return active


def _bulk_verify_status_counts(job_id: int, owner_id: str):
    counts = {"ok": 0, "mismatch": 0, "error": 0, "skipped": 0, "processed": 0}
    items = BulkCalcVerifyJobItem.query.filter_by(job_id=job_id, owner_id=owner_id).all()
    for item in items:
        status = _norm_text(item.status)
        if status in counts:
            counts[status] += 1
        if status in {"ok", "mismatch", "error", "skipped"}:
            counts["processed"] += 1
    return counts


def _bulk_verify_refresh_job_counters(job: BulkCalcVerifyJob):
    counts = _bulk_verify_status_counts(job.id, job.owner_id)
    job.processed_count = counts["processed"]
    job.ok_count = counts["ok"]
    job.mismatch_count = counts["mismatch"]
    job.error_count = counts["error"]
    job.skipped_count = counts["skipped"]
    if counts["processed"] >= int(job.total_count or 0):
        job.status = "completed"
        if not job.finished_at:
            job.finished_at = datetime.utcnow()


def _bulk_verify_item_payload(item: BulkCalcVerifyJobItem):
    return {
        "uid": _norm_text(item.account_uid),
        "account_uid": _norm_text(item.account_uid),
        "status": _norm_text(item.status),
        "reason": _norm_text(item.reason) or _norm_text(item.error_message),
        "old_summary": _bulk_verify_json_load(item.old_summary_json),
        "new_summary": _bulk_verify_json_load(item.new_summary_json),
        "diff": _bulk_verify_json_load(item.diff_json),
        "error_code": _norm_text(item.error_code),
    }


def _bulk_verify_job_response(job: BulkCalcVerifyJob):
    items = BulkCalcVerifyJobItem.query.filter_by(job_id=job.id, owner_id=job.owner_id).order_by(BulkCalcVerifyJobItem.id.asc()).all()
    return {
        "success": True,
        "job_id": int(job.id),
        "status": _norm_text(job.status) or "queued",
        "total": int(job.total_count or 0),
        "processed": int(job.processed_count or 0),
        "ok_count": int(job.ok_count or 0),
        "ok_items": int(job.ok_count or 0),
        "ok": int(job.ok_count or 0),
        "mismatch": int(job.mismatch_count or 0),
        "error": int(job.error_count or 0),
        "skipped": int(job.skipped_count or 0),
        "reason": _norm_text(job.reason),
        "message": _norm_text(job.error_message),
        "items": [_bulk_verify_item_payload(item) for item in items],
    }


def _bulk_verify_create_job(owner_id: str, user_id: str, raw_uids, reason: str):
    requested = _recalc_normalize_uids(raw_uids)
    if not requested:
        return None, {"requested": 0}
    if len(requested) > RECALC_BATCH_MAX_UIDS:
        return "TOO_MANY_UIDS", {"max_uids": RECALC_BATCH_MAX_UIDS, "requested": len(requested)}

    app.logger.info("[stage16][bulk-verify] start owner_id=%s requested=%s", owner_id, len(requested))
    targets = _owner_abonent_summary_targets(owner_id)
    targets_by_uid = {_norm_text(t.get("account_uid")): t for t in targets if _norm_text(t.get("account_uid"))}
    active_uids = _bulk_verify_active_uids(owner_id)
    reason_norm = _norm_text(reason) or "STAGE16_BULK_VERIFY"

    job = BulkCalcVerifyJob(owner_id=owner_id, requested_by=user_id, reason=reason_norm, status="queued", total_count=len(requested))
    db.session.add(job)
    db.session.flush()

    for uid in requested:
        if uid not in targets_by_uid:
            db.session.add(BulkCalcVerifyJobItem(
                job_id=job.id,
                owner_id=owner_id,
                account_uid=uid,
                status="skipped",
                reason="UID_NOT_FOUND",
                error_code="UID_NOT_FOUND",
                finished_at=datetime.utcnow(),
            ))
            continue
        if uid in active_uids:
            app.logger.info("[stage16][bulk-verify] already_running owner_id=%s uid=%s", owner_id, uid)
            db.session.add(BulkCalcVerifyJobItem(
                job_id=job.id,
                owner_id=owner_id,
                account_uid=uid,
                status="skipped",
                reason="already_running",
                error_code="already_running",
                finished_at=datetime.utcnow(),
            ))
            continue
        db.session.add(BulkCalcVerifyJobItem(job_id=job.id, owner_id=owner_id, account_uid=uid, status="queued"))

    db.session.flush()
    _bulk_verify_refresh_job_counters(job)
    db.session.commit()
    return job, {"requested": len(requested)}


def _bulk_verify_process_job(owner_id: str, job: BulkCalcVerifyJob, step_limit: int = 25):
    if not job or job.owner_id != owner_id or job.status not in {"queued", "running"}:
        return
    if job.status == "queued":
        job.status = "running"
        if not job.started_at:
            job.started_at = datetime.utcnow()
        db.session.commit()

    items = (
        BulkCalcVerifyJobItem.query
        .filter_by(job_id=job.id, owner_id=owner_id)
        .filter(BulkCalcVerifyJobItem.status.in_(["queued", "running"]))
        .order_by(BulkCalcVerifyJobItem.id.asc())
        .limit(max(1, int(step_limit)))
        .all()
    )
    for item in items:
        item.status = "running"
        if not item.started_at:
            item.started_at = datetime.utcnow()
        db.session.commit()
        try:
            summary_row = AbonentSummary.query.filter_by(owner_id=owner_id, account_uid=item.account_uid).order_by(AbonentSummary.id.asc()).first()
            snapshot_row = CardSnapshot.query.filter_by(owner_id=owner_id, abonent_uid=item.account_uid).first()
            old_summary = _bulk_verify_json_load(summary_row.summary_json if summary_row else "{}", {})
            if not snapshot_row:
                raise ValueError("CARD_SNAPSHOT_MISSING")
            snapshot = _bulk_verify_json_load(snapshot_row.snapshot_json, {})
            new_summary = _bulk_verify_summary_from_snapshot(snapshot)
            new_status = _summary_status_from_payload(new_summary)
            if new_status in {"error", "invalid"}:
                raise ValueError(_norm_text(new_summary.get("summary_reason")) or "CARD_SNAPSHOT_ERROR")
            old_view, new_view, diff = _bulk_verify_diff(old_summary, new_summary)
            item.old_summary_json = json.dumps(old_view, ensure_ascii=False, sort_keys=True)
            item.new_summary_json = json.dumps(new_view, ensure_ascii=False, sort_keys=True)
            item.diff_json = json.dumps(diff, ensure_ascii=False, sort_keys=True)
            if diff:
                item.status = "mismatch"
                item.reason = "SUMMARY_SNAPSHOT_MISMATCH"
                app.logger.info("[stage16][bulk-verify] item mismatch owner_id=%s uid=%s", owner_id, item.account_uid)
            else:
                item.status = "ok"
                item.reason = "OK"
                app.logger.info("[stage16][bulk-verify] item ok owner_id=%s uid=%s", owner_id, item.account_uid)
        except Exception as exc:
            code = _norm_text(str(exc)) or "BULK_VERIFY_ITEM_FAILED"
            item.status = "error"
            item.reason = code
            item.error_code = code
            item.error_message = code
            app.logger.info("[stage16][bulk-verify] item error owner_id=%s uid=%s error=%s", owner_id, item.account_uid, code)
        item.finished_at = datetime.utcnow()
        db.session.commit()

    _bulk_verify_refresh_job_counters(job)
    if job.status == "completed":
        app.logger.info("[stage16][bulk-verify] completed owner_id=%s job_id=%s", owner_id, job.id)
    db.session.commit()



def _parse_summary_status_filter(raw_value: str):
    raw = _norm_text(raw_value).lower()
    if not raw:
        return None
    parts = [x.strip() for x in re.split(r"[,\s]+", raw) if x.strip()]
    statuses = set()
    for part in parts:
        if part in {"stale", "устаревшие", "устаревший"}:
            statuses.update({"dirty", "missing", "error"})
        elif part in {"fresh", "dirty", "missing", "error"}:
            statuses.add(part)
    return statuses if statuses else None


def _target_matches_abonent_query(target: dict, query_text: str):
    q = _norm_text(query_text).lower()
    if not q:
        return True
    identity = target.get("identity") if isinstance(target.get("identity"), dict) else {}
    values = [
        target.get("abonent_id"),
        target.get("account_uid"),
        target.get("account_number"),
        identity.get("fio"),
        identity.get("address"),
    ]
    haystack = " ".join(_norm_text(value).lower() for value in values)
    tokens = [x for x in re.split(r"\s+", q) if x]
    return all(token in haystack for token in tokens)


def _abonent_index_payload(target: dict, summary: dict):
    identity = target.get("identity") if isinstance(target.get("identity"), dict) else {}
    summary_payload = _summary_without_stale_totals(summary)
    if "abonent" not in summary_payload or not isinstance(summary_payload.get("abonent"), dict):
        summary_payload["abonent"] = {
            "abonent_id": _norm_text(identity.get("abonent_id") or target.get("abonent_id")),
            "account_uid": _norm_text(identity.get("account_uid") or target.get("account_uid")),
            "account_number": _norm_text(identity.get("account_number") or target.get("account_number")),
            "fio": _norm_text(identity.get("fio")),
            "address": _norm_text(identity.get("address")),
        }
    if not _norm_text(summary_payload.get("fio")) and _norm_text(identity.get("fio")):
        summary_payload["fio"] = _norm_text(identity.get("fio"))
    if not _norm_text(summary_payload.get("address")) and _norm_text(identity.get("address")):
        summary_payload["address"] = _norm_text(identity.get("address"))
    return {
        "abonent_id": _norm_text(target.get("abonent_id")),
        "account_uid": _norm_text(target.get("account_uid")),
        "account_number": _norm_text(target.get("account_number")),
        "summary_status": _summary_status_from_payload(summary_payload),
        "summary": summary_payload,
    }


def _abonents_table_name():
    for table_name in ("abonent", "abonents"):
        if _table_columns(table_name):
            return table_name
    return None


def _abonents_api_select_config():
    abonent_table = _abonents_table_name()
    if not abonent_table:
        return None

    abonent_cols = _table_columns(abonent_table)
    premise_cols = _table_columns("premise")
    summary_cols = _table_columns("abonent_summary")
    if not summary_cols:
        return None

    uid_col = _first_existing_column(abonent_cols, ("uid", "account_uid", "abonent_uid"))
    abonent_id_col = _first_existing_column(abonent_cols, ("abonent_id", "id"))
    account_col = _first_existing_column(abonent_cols, ("account_number", "ls", "personal_account", "account", "id"))
    fio_col = _first_existing_column(abonent_cols, ("fio", "full_name", "fullName", "name"))
    owner_col = _first_existing_column(abonent_cols, ("owner_id", "owner"))
    if not owner_col or not uid_col:
        return None

    premise_join = ""
    if premise_cols:
        premise_id_col = _first_existing_column(abonent_cols, ("premise_id", "premise", "premise_uid"))
        premise_pk_col = _first_existing_column(premise_cols, ("id", "premise_id", "uid"))
        premise_owner_col = _first_existing_column(premise_cols, ("owner_id", "owner"))
        premise_abonent_uid_col = _first_existing_column(premise_cols, ("abonent_uid", "account_uid", "uid"))
        premise_abonent_id_col = _first_existing_column(premise_cols, ("abonent_id",))
        if premise_id_col and premise_pk_col:
            premise_join = f"LEFT JOIN premise p ON p.{_sql_ident(premise_pk_col)} = a.{_sql_ident(premise_id_col)}"
        elif premise_abonent_uid_col:
            premise_join = f"LEFT JOIN premise p ON p.{_sql_ident(premise_abonent_uid_col)} = a.{_sql_ident(uid_col)}"
        elif premise_abonent_id_col and abonent_id_col:
            premise_join = f"LEFT JOIN premise p ON p.{_sql_ident(premise_abonent_id_col)} = a.{_sql_ident(abonent_id_col)}"
        if premise_join and premise_owner_col:
            premise_join += f" AND p.{_sql_ident(premise_owner_col)} = :owner"
        if not premise_join:
            premise_cols = set()

    regnum_col = _first_existing_column(premise_cols, ("regnum", "registration_number"))
    premise_address_col = _first_existing_column(premise_cols, ("address", "full_address", "addr"))
    abonent_address_col = _first_existing_column(abonent_cols, ("address", "full_address", "addr"))

    return {
        "abonent_table": abonent_table,
        "abonent_cols": abonent_cols,
        "premise_cols": premise_cols,
        "uid_col": uid_col,
        "abonent_id_col": abonent_id_col,
        "account_col": account_col,
        "fio_col": fio_col,
        "owner_col": owner_col,
        "premise_join": premise_join,
        "regnum_col": regnum_col,
        "premise_address_col": premise_address_col,
        "abonent_address_col": abonent_address_col,
    }


def _abonents_api_row_payload(row):
    data = dict(row._mapping)
    has_summary = data.get("summary_row_id") is not None
    summary_status = _norm_text(data.get("summary_status")) if has_summary else "missing"
    summary_reason = _norm_text(data.get("summary_reason")) if has_summary else "SUMMARY_NOT_BUILT"
    snapshot_status = _cache_status(data.get("snapshot_status") if data.get("snapshot_row_id") is not None else "missing")
    snapshot_reason = _cache_reason("snapshot", snapshot_status, data.get("snapshot_reason") if data.get("snapshot_row_id") is not None else "")
    summary_hash = _norm_text(data.get("summary_input_hash")) if has_summary else ""
    snapshot_hash = _norm_text(data.get("snapshot_input_hash")) if data.get("snapshot_row_id") is not None else ""
    hash_mismatch = summary_hash != snapshot_hash
    warnings = []
    if hash_mismatch:
        warnings.append("INPUT_HASH_CHANGED")
    if summary_status == "fresh" and data.get("snapshot_row_id") is None:
        warnings.append("SUMMARY_FRESH_CARD_SNAPSHOT_MISSING")
    if summary_status == "fresh" and snapshot_status in {"dirty", "error", "invalid"}:
        warnings.append("SUMMARY_FRESH_SNAPSHOT_NOT_FRESH")

    payload = {
        "abonent_id": _norm_text(data.get("abonent_id")),
        "account_number": _norm_text(data.get("account_number")),
        "uid": _norm_text(data.get("uid")),
        "fio": _norm_text(data.get("fio")),
        "premise": _norm_text(data.get("premise")),
        "regnum": _norm_text(data.get("regnum")),
        "address": _norm_text(data.get("address")),
        "summary_status": summary_status or "missing",
        "summary_reason": summary_reason,
        "total_debt": _decimal_json_or_none(data.get("total_debt")) if has_summary else None,
        "total_accrued": _decimal_json_or_none(data.get("total_accrued")) if has_summary else None,
        "total_paid": _decimal_json_or_none(data.get("total_paid")) if has_summary else None,
        "penalty_debt": _decimal_json_or_none(data.get("total_penalty")) if has_summary else None,
        "total_penalty": _decimal_json_or_none(data.get("total_penalty")) if has_summary else None,
        "updated_at": _dt_json_or_none(data.get("updated_at")) if has_summary else None,
        "snapshot_status": snapshot_status,
        "snapshot_reason": snapshot_reason,
        "input_hash_summary": summary_hash,
        "input_hash_snapshot": snapshot_hash,
        "hash_mismatch": hash_mismatch,
        "warnings": sorted(set(warnings)),
    }
    for key, value in data.items():
        if key.startswith("address_"):
            payload[key] = _norm_text(value)
    return payload


def _abonents_api_query(owner: str, search_text: str, status_filter=None):
    cfg = _abonents_api_select_config()
    if cfg is None:
        raise RuntimeError("abonents_table_missing")

    a_table = _sql_ident(cfg["abonent_table"])
    owner_expr = f"a.{_sql_ident(cfg['owner_col'])}"
    uid_expr = f"a.{_sql_ident(cfg['uid_col'])}"
    abonent_id_expr = f"a.{_sql_ident(cfg['abonent_id_col'])}" if cfg["abonent_id_col"] else "''"
    account_expr = f"a.{_sql_ident(cfg['account_col'])}" if cfg["account_col"] else abonent_id_expr
    fio_expr = f"a.{_sql_ident(cfg['fio_col'])}" if cfg["fio_col"] else "''"
    premise_expr = "p.`id`" if "id" in cfg["premise_cols"] else "''"
    regnum_expr = f"p.{_sql_ident(cfg['regnum_col'])}" if cfg["regnum_col"] else "''"
    premise_address_expr = f"p.{_sql_ident(cfg['premise_address_col'])}" if cfg["premise_address_col"] else "''"
    abonent_address_expr = f"a.{_sql_ident(cfg['abonent_address_col'])}" if cfg["abonent_address_col"] else "''"

    address_selects = []
    for col in ("city", "street", "house", "building", "corpus", "flat", "apartment", "room"):
        if col in cfg["premise_cols"]:
            address_selects.append(f"p.{_sql_ident(col)} AS address_{col}")

    where_parts = [f"{owner_expr} = :owner"]
    params = {"owner": owner}
    q = _norm_text(search_text).lower()
    if q:
        search_exprs = [abonent_id_expr, account_expr, uid_expr, fio_expr, premise_address_expr, regnum_expr]
        tokens = [x for x in re.split(r"\s+", q) if x]
        for index, token in enumerate(tokens):
            key = f"search_{index}"
            params[key] = f"%{token}%"
            like_parts = [
                f"LOWER(COALESCE({_sql_cast_text(expr)}, '')) LIKE :{key}"
                for expr in search_exprs
            ]
            where_parts.append("(" + " OR ".join(like_parts) + ")")
    if status_filter:
        status_names = []
        for value in sorted(status_filter):
            clean = _norm_text(value).lower()
            if clean in {"fresh", "dirty", "missing", "error", "invalid"}:
                status_names.append(clean)
        if status_names:
            placeholders = []
            for index, status_name in enumerate(status_names):
                key = f"status_{index}"
                params[key] = status_name
                placeholders.append(f":{key}")
            where_parts.append("LOWER(COALESCE(s.summary_status, 'missing')) IN (" + ", ".join(placeholders) + ")")

    select_sql = ",\n            ".join([
        f"{abonent_id_expr} AS abonent_id",
        f"{account_expr} AS account_number",
        f"{uid_expr} AS uid",
        f"{fio_expr} AS fio",
        f"{premise_expr} AS premise",
        f"{regnum_expr} AS regnum",
        f"COALESCE(NULLIF({_sql_cast_text(premise_address_expr)}, ''), NULLIF({_sql_cast_text(abonent_address_expr)}, '')) AS address",
        "s.id AS summary_row_id",
        "s.summary_status AS summary_status",
        "s.summary_reason AS summary_reason",
        "s.input_hash AS summary_input_hash",
        "s.total_debt AS total_debt",
        "s.total_accrued AS total_accrued",
        "s.total_paid AS total_paid",
        "s.penalty_debt AS total_penalty",
        "s.updated_at AS updated_at",
        "c.id AS snapshot_row_id",
        "c.snapshot_status AS snapshot_status",
        "c.snapshot_reason AS snapshot_reason",
        "c.input_hash AS snapshot_input_hash",
        *address_selects,
    ])
    from_sql = f"""
        FROM {a_table} a
        {cfg["premise_join"]}
        LEFT JOIN abonent_summary s
            ON s.id = (
                SELECT s2.id
                FROM abonent_summary s2
                WHERE s2.owner_id = :owner
                    AND s2.account_uid = {uid_expr}
                ORDER BY s2.updated_at DESC, s2.id DESC
                LIMIT 1
            )
        LEFT JOIN card_snapshot c
            ON c.owner_id = :owner
            AND c.abonent_uid = {uid_expr}
    """
    where_sql = " AND ".join(where_parts)
    order_sql = f"ORDER BY {fio_expr}, {account_expr}, {uid_expr}"
    return select_sql, from_sql, where_sql, order_sql, params


def _legacy_abonents_index_response(owner: str, page: int, per_page: int, query_text: str):
    status_filter = _parse_summary_status_filter(
        request.args.get("summary_status") or request.args.get("status") or ""
    )

    targets = [t for t in _owner_abonent_summary_targets(owner) if _target_matches_abonent_query(t, query_text)]
    start = (page - 1) * per_page

    def load_summaries_by_uid(page_targets):
        uids = [_norm_text(t.get("account_uid")) for t in page_targets if _norm_text(t.get("account_uid"))]
        if not uids:
            return {}
        summary_rows = (
            AbonentSummary.query
            .filter_by(owner_id=owner)
            .filter(AbonentSummary.account_uid.in_(uids))
            .order_by(AbonentSummary.updated_at.desc(), AbonentSummary.id.desc())
            .all()
        )
        summaries = {}
        for row in summary_rows:
            uid = _norm_text(row.account_uid)
            if uid and uid not in summaries:
                summaries[uid] = row
        return summaries

    if status_filter is None:
        total = len(targets)
        page_targets = targets[start:start + per_page]
        summaries_by_uid = load_summaries_by_uid(page_targets)
        page_pairs = [
            (target, _summary_from_row_or_missing(summaries_by_uid.get(_norm_text(target.get("account_uid"))), target))
            for target in page_targets
        ]
    else:
        summaries_by_uid = load_summaries_by_uid(targets)
        filtered = []
        for target in targets:
            uid = _norm_text(target.get("account_uid"))
            summary = _summary_from_row_or_missing(summaries_by_uid.get(uid), target)
            status = _summary_status_from_payload(summary)
            if status in status_filter:
                filtered.append((target, summary))

        total = len(filtered)
        page_pairs = filtered[start:start + per_page]

    return jsonify(
        ok=True,
        items=[_abonent_index_payload(target, summary) for target, summary in page_pairs],
        pagination=_pagination_payload(page, per_page, total),
    )


def _row_human_error_payload(r: ImportBatchRow):
    return {
        "excel_row_ref": r.excel_row_ref,
        "excel_sheet_name": r.excel_sheet_name,
        "account_uid": r.account_uid,
        "account_number": r.account_number,
        "payment_date": r.payment_date,
        "payment_period": r.payment_period,
        "amount": r.amount,
        "source_index": r.source_index,
        "source_label": r.source_label,
        "reason_code": r.reason_code,
        "reason_text": r.reason_text,
        "recommendation": json.loads(r.error_details_json or "{}").get("recommendation", ""),
    }


def _import_batches_schema_status():
    try:
        rows = db.session.execute(text("DESCRIBE import_batches")).all()
    except SQLAlchemyError as exc:
        db.session.rollback()
        app.logger.error(
            "Import DB schema check failed for import_batches. Run initdb/migrations before import. error=%s",
            exc,
        )
        return {
            "ok": False,
            "error": "import_batches_schema_check_failed",
            "missing_columns": list(IMPORT_BATCH_CRITICAL_COLUMNS),
            "migration_sql": IMPORT_BATCH_AUDIT_FIELDS_MIGRATION_SQL,
        }

    existing = {row[0] for row in rows}
    missing = [col for col in IMPORT_BATCH_CRITICAL_COLUMNS if col not in existing]
    if missing:
        app.logger.error(
            "Import DB schema mismatch: import_batches is missing critical columns %s. Required migration: %s",
            ", ".join(missing),
            IMPORT_BATCH_AUDIT_FIELDS_MIGRATION_SQL,
        )
    return {
        "ok": not missing,
        "error": "import_batches_missing_columns" if missing else "",
        "missing_columns": missing,
        "migration_sql": IMPORT_BATCH_AUDIT_FIELDS_MIGRATION_SQL if missing else "",
    }


def _import_schema_error_response():
    schema = _import_batches_schema_status()
    if schema["ok"]:
        return None
    return jsonify(
        ok=False,
        error=schema["error"],
        details={
            "table": "import_batches",
            "missing_columns": schema["missing_columns"],
            "migration_sql": schema["migration_sql"],
        },
    ), 503


@app.before_request
def _guard_import_batches_schema():
    if request.path.startswith("/api/import"):
        return _import_schema_error_response()
    return None


@app.get("/health")
def health():
    schema = _import_batches_schema_status()
    status = "ok" if schema["ok"] else "degraded"
    code = 200 if schema["ok"] else 503
    return jsonify(
        status=status,
        checks={
            "import_batches_schema": {
                "ok": schema["ok"],
                "missing_columns": schema["missing_columns"],
            },
        },
    ), code


@app.get("/")
def index():
    return "JKH API: ok\n"


@app.post("/api/admin/initdb")
def initdb():
    db.create_all()
    return jsonify(ok=True, message="users + kv_store ready")


@app.get("/api/abonents")
def abonents_index_list():
    # CRITICAL GUARD: lightweight index endpoint is read-only DB summary transport.
    # Owner is always taken from session; never trust owner/query owner from client.
    # Keep this handler list-only: no ledger reads, no calculation side effects,
    # and no synthesized totals when abonent_summary is missing or errored.
    user, err = _require_user()
    if err:
        return err

    owner = user.id
    page, limit = _parse_pagination_args(default_per_page=50, max_per_page=200)
    query_text = request.args.get("query", request.args.get("search", ""))

    try:
        status_filter = _parse_summary_status_filter(
            request.args.get("summary_status") or request.args.get("status") or ""
        )
        select_sql, from_sql, where_sql, order_sql, params = _abonents_api_query(owner, query_text, status_filter)
        if select_sql is None:
            return _legacy_abonents_index_response(owner, page, limit, query_text)
        total = db.session.execute(
            text(f"SELECT COUNT(*) AS total {from_sql} WHERE {where_sql}"),
            params,
        ).scalar() or 0
        rows = db.session.execute(
            text(
                f"""
                SELECT {select_sql}
                {from_sql}
                WHERE {where_sql}
                {order_sql}
                LIMIT :limit OFFSET :offset
                """
            ),
            {**params, "limit": limit, "offset": (page - 1) * limit},
        ).all()
    except RuntimeError as exc:
        if str(exc) == "abonents_table_missing":
            return _legacy_abonents_index_response(owner, page, limit, query_text)
        raise
    except SQLAlchemyError as exc:
        db.session.rollback()
        app.logger.exception("GET /api/abonents failed for owner=%s", owner)
        return jsonify(ok=False, error="abonents_query_failed", details=str(exc)), 500

    return jsonify(
        ok=True,
        items=[_abonents_api_row_payload(row) for row in rows],
        page=page,
        limit=limit,
        total=total,
    )


@app.get("/api/abonent_summary")
def abonent_summary_list():
    # CRITICAL GUARD: Summary API is read-only derived-cache transport.
    # Do not add recalculation, ledger rebuild, autoaccrual, fallback totals,
    # hidden writes, implicit refresh, repair of missing summary rows, or
    # payments_<uid> reads inside GET /api/abonent_summary.
    user, err = _require_user()
    if err:
        return err

    owner = request.args.get("owner") if user.role == "admin" else user.id
    page, per_page = _parse_pagination_args()

    q = AbonentSummary.query
    if owner:
        q = q.filter_by(owner_id=owner)

    abonent_id = str(request.args.get("abonent_id") or "").strip()
    account_uid = str(request.args.get("account_uid") or "").strip()
    account_number = str(request.args.get("account_number") or "").strip()

    if abonent_id:
        q = q.filter_by(abonent_id=abonent_id)
    if account_uid:
        q = q.filter_by(account_uid=account_uid)
    if account_number:
        q = q.filter_by(account_number=account_number)

    total = q.count()
    items = (
        q.order_by(AbonentSummary.updated_at.desc(), AbonentSummary.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return jsonify(
        ok=True,
        items=[_abonent_summary_payload(x) for x in items],
        pagination=_pagination_payload(page, per_page, total),
    )


@app.post("/api/abonent_summary/mark_dirty")
def abonent_summary_mark_dirty():
    user, err = _require_user()
    if err:
        return err

    owner = user.id
    counters = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify(ok=False, error="summary_invalid", counters=counters), 400

    account_uid = _norm_text(body.get("account_uid"))
    if not account_uid:
        return jsonify(ok=False, error="account_uid_required", counters=counters), 400

    reason = _norm_text(body.get("reason")) or "UNKNOWN_CHANGE"
    if reason not in ABONENT_SUMMARY_DIRTY_REASONS:
        return jsonify(ok=False, error="invalid_reason", counters=counters), 400

    try:
        targets = _owner_abonent_summary_targets(owner)
        target = next((t for t in targets if _norm_text(t.get("account_uid")) == account_uid), None)
        if not target:
            return jsonify(ok=False, error="uid_not_found"), 404

        if reason == "CALC_PERIOD_CHANGED":
            counters["skipped"] += 1
            return jsonify(
                ok=True,
                account_uid=account_uid,
                status="skipped",
                reason=reason,
                view_only_reason=reason,
                counters=counters,
            )

        row = AbonentSummary.query.filter_by(owner_id=owner, account_uid=account_uid).order_by(AbonentSummary.id.asc()).first()
        if row:
            try:
                existing_summary = json.loads(row.summary_json or "{}")
            except (TypeError, ValueError):
                existing_summary = {}
            row.abonent_id = _norm_text(target.get("abonent_id"))
            row.account_number = _norm_text(target.get("account_number"))
            dirty_summary = _dirty_abonent_summary_payload(existing_summary, reason)
            _set_abonent_summary_json(row, target, dirty_summary)
            counters["updated"] += 1
        else:
            dirty_summary = _dirty_abonent_summary_payload(None, reason)
            row = AbonentSummary(
                owner_id=owner,
                abonent_id=_norm_text(target.get("abonent_id")),
                account_uid=account_uid,
                account_number=_norm_text(target.get("account_number")),
            )
            _set_abonent_summary_json(row, target, dirty_summary)
            db.session.add(row)
            counters["created"] += 1

        db.session.commit()
    except SQLAlchemyError as exc:
        db.session.rollback()
        app.logger.exception("abonent_summary mark_dirty failed for owner=%s", owner)
        return jsonify(ok=False, error="summary_mark_dirty_failed", counters=counters, details=str(exc)), 500

    return jsonify(
        ok=True,
        account_uid=account_uid,
        status="dirty",
        reason=reason,
        counters=counters,
    )


@app.post("/api/abonent_summary/recalc_batch")
def abonent_summary_recalc_batch():
    user, err = _require_user()
    if err:
        return err

    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify(ok=False, error="summary_invalid"), 400

    raw_uids = body.get("account_uids")
    if raw_uids is None:
        raw_uids = body.get("uids")
    if raw_uids is None:
        raw_uids = body.get("account_uid")
    if isinstance(raw_uids, str):
        raw_uids = [raw_uids]
    if not isinstance(raw_uids, list):
        return jsonify(ok=False, error="account_uids_required"), 400

    requested = []
    seen = set()
    for value in raw_uids:
        uid = _norm_text(value)
        if not uid or uid in seen:
            continue
        seen.add(uid)
        requested.append(uid)

    if not requested:
        return jsonify(ok=False, error="account_uids_required"), 400

    targets = _owner_abonent_summary_targets(user.id)
    targets_by_uid = {_norm_text(t.get("account_uid")): t for t in targets if _norm_text(t.get("account_uid"))}

    items = []
    for uid in requested:
        target = targets_by_uid.get(uid)
        items.append({
            "account_uid": uid,
            "allowed": bool(target),
            "status": "allowed" if target else "not_found",
        })

    return jsonify(ok=True, allowed_uids=[i["account_uid"] for i in items if i["allowed"]], items=items)


@app.post("/api/abonent_summary/recalc_batch_job")
def abonent_summary_recalc_batch_job_create():
    user, err = _require_user()
    if err:
        return err
    body = request.get_json(silent=True) or {}
    raw_uids = body.get("uids")
    if raw_uids is None:
        raw_uids = body.get("account_uids")
    if isinstance(raw_uids, str):
        raw_uids = [raw_uids]
    if not isinstance(raw_uids, list):
        return jsonify(ok=False, error="account_uids_required"), 400
    job, counters = _recalc_batch_create_job(user.id, user.id, raw_uids, body.get("reason"))
    if job == "TOO_MANY_UIDS":
        return jsonify(ok=False, error="TOO_MANY_UIDS", details={"max_uids": counters.get("max_uids"), "requested": counters.get("requested")}), 400
    if not job:
        return jsonify(ok=False, error="account_uids_required"), 400
    return jsonify(ok=True, job_id=int(job.id), status="queued", **counters)


@app.post("/api/abonent_summary/recalc_batch_job/<int:job_id>/run")
def abonent_summary_recalc_batch_job_run(job_id: int):
    user, err = _require_user()
    if err:
        return err
    job = RecalcBatchJob.query.filter_by(id=job_id, owner_id=user.id).first()
    if not job:
        return jsonify(ok=False, error="job_not_found"), 404
    if job.status in {"completed", "failed", "stale"}:
        return jsonify(**_batch_job_status_response(job))
    _recalc_batch_process_job(user.id, job, step_limit=1)
    db.session.refresh(job)
    return jsonify(**_batch_job_status_response(job))


@app.get("/api/abonent_summary/recalc_batch_job/latest")
def abonent_summary_recalc_batch_job_latest():
    user, err = _require_user()
    if err:
        return err
    job = RecalcBatchJob.query.filter_by(owner_id=user.id).order_by(RecalcBatchJob.id.desc()).first()
    if not job:
        return jsonify(ok=True, job=None)
    return jsonify(ok=True, job=_batch_job_payload(job))


@app.get("/api/abonent_summary/recalc_batch_job/<int:job_id>")
def abonent_summary_recalc_batch_job_status(job_id: int):
    user, err = _require_user()
    if err:
        return err
    job = RecalcBatchJob.query.filter_by(id=job_id, owner_id=user.id).first()
    if not job:
        return jsonify(ok=False, error="job_not_found"), 404
    _recalc_batch_process_job(user.id, job, step_limit=10)
    db.session.refresh(job)
    return jsonify(**_batch_job_status_response(job))


@app.post("/api/abonent_summary/bulk_calc_verify")
def abonent_summary_bulk_calc_verify_create():
    user, err = _require_user()
    if err:
        return err
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify(success=False, error="summary_invalid"), 400
    raw_uids = body.get("uids")
    if isinstance(raw_uids, str):
        raw_uids = [raw_uids]
    if not isinstance(raw_uids, list):
        return jsonify(success=False, error="uids_required"), 400
    job, counters = _bulk_verify_create_job(user.id, user.id, raw_uids, body.get("reason"))
    if job == "TOO_MANY_UIDS":
        return jsonify(success=False, error="TOO_MANY_UIDS", details={"max_uids": counters.get("max_uids"), "requested": counters.get("requested")}), 400
    if not job:
        return jsonify(success=False, error="uids_required"), 400
    db.session.refresh(job)
    return jsonify(**_bulk_verify_job_response(job))


@app.get("/api/abonent_summary/bulk_calc_verify/<int:job_id>")
def abonent_summary_bulk_calc_verify_status(job_id: int):
    user, err = _require_user()
    if err:
        return err
    job = BulkCalcVerifyJob.query.filter_by(id=job_id, owner_id=user.id).first()
    if not job:
        return jsonify(success=False, error="job_not_found"), 404
    _bulk_verify_process_job(user.id, job, step_limit=10)
    db.session.refresh(job)
    return jsonify(**_bulk_verify_job_response(job))


@app.post("/api/abonent_summary/rebuild")
def abonent_summary_rebuild():
    user, err = _require_user()
    if err:
        return err

    owner = user.id
    counters = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}
    body = request.get_json(silent=True)

    try:
        if body:
            if not isinstance(body, dict):
                return jsonify(ok=False, error="summary_invalid", counters=counters), 400

            account_uid = _norm_text(body.get("account_uid"))
            if not account_uid:
                return jsonify(ok=False, error="account_uid_required", counters=counters), 400

            summary = body.get("summary")
            if not isinstance(summary, dict):
                return jsonify(ok=False, error="summary_invalid", counters=counters), 400
            summary_scope = _norm_text(summary.get("summary_scope") or summary.get("report_scope")).lower()
            if summary_scope in {"period", "report"}:
                return jsonify(ok=False, error="period_summary_not_allowed", counters=counters), 400

            targets = _owner_abonent_summary_targets(owner)
            target = next((t for t in targets if _norm_text(t.get("account_uid")) == account_uid), None)
            if not target:
                return jsonify(ok=False, error="uid_not_found", counters=counters), 404
            summary = _summary_without_stale_totals(summary, target, owner)

            abonent_id = _norm_text(body.get("abonent_id")) or _norm_text(target.get("abonent_id"))
            account_number = _norm_text(body.get("account_number")) or _norm_text(target.get("account_number"))
            row = AbonentSummary.query.filter_by(owner_id=owner, account_uid=account_uid).order_by(AbonentSummary.id.asc()).first()
            if row:
                row.abonent_id = abonent_id
                row.account_number = account_number
                _set_abonent_summary_json(row, target, summary)
                counters["updated"] += 1
            else:
                row = AbonentSummary(
                    owner_id=owner,
                    abonent_id=abonent_id,
                    account_uid=account_uid,
                    account_number=account_number,
                )
                _set_abonent_summary_json(row, target, summary)
                db.session.add(row)
                counters["created"] += 1

            db.session.commit()
            return jsonify(
                ok=True,
                counters=counters,
                summary_status=_summary_status_from_payload(summary),
                summary_reason=_norm_text(summary.get("summary_reason") or summary.get("reason")),
            )

        targets = _owner_abonent_summary_targets(owner)
        for target in targets:
            account_uid = _norm_text(target.get("account_uid"))
            if not account_uid:
                counters["skipped"] += 1
                continue

            missing_summary = _build_missing_abonent_summary(target)
            row = AbonentSummary.query.filter_by(owner_id=owner, account_uid=account_uid).order_by(AbonentSummary.id.asc()).first()
            if row:
                row.abonent_id = _norm_text(target.get("abonent_id"))
                row.account_number = _norm_text(target.get("account_number"))
                _set_abonent_summary_json(row, target, missing_summary)
                counters["updated"] += 1
            else:
                row = AbonentSummary(
                    owner_id=owner,
                    abonent_id=_norm_text(target.get("abonent_id")),
                    account_uid=account_uid,
                    account_number=_norm_text(target.get("account_number")),
                )
                _set_abonent_summary_json(row, target, missing_summary)
                db.session.add(row)
                counters["created"] += 1

        db.session.commit()
    except SQLAlchemyError as exc:
        db.session.rollback()
        app.logger.exception("abonent_summary rebuild failed for owner=%s", owner)
        return jsonify(ok=False, error="summary_rebuild_failed", counters=counters, details=str(exc)), 500

    return jsonify(ok=True, counters=counters)


def _audit_snapshot_summary_targets(owner_id: str = ""):
    owner_id = _norm_text(owner_id)
    cfg = _abonents_api_select_config()
    if cfg is None:
        owners = [owner_id] if owner_id else [_norm_text(row[0]) for row in db.session.query(User.id).all()]
        targets = []
        for owner in owners:
            for target in _owner_abonent_summary_targets(owner):
                target = dict(target)
                target["owner_id"] = owner
                targets.append(target)
        return targets

    a_table = _sql_ident(cfg["abonent_table"])
    owner_expr = f"a.{_sql_ident(cfg['owner_col'])}"
    uid_expr = f"a.{_sql_ident(cfg['uid_col'])}"
    abonent_id_expr = f"a.{_sql_ident(cfg['abonent_id_col'])}" if cfg["abonent_id_col"] else "''"
    account_expr = f"a.{_sql_ident(cfg['account_col'])}" if cfg["account_col"] else abonent_id_expr
    fio_expr = f"a.{_sql_ident(cfg['fio_col'])}" if cfg["fio_col"] else "''"
    address_expr = f"a.{_sql_ident(cfg['abonent_address_col'])}" if cfg["abonent_address_col"] else "''"
    where_sql = f"WHERE {owner_expr} = :owner" if owner_id else ""
    params = {"owner": owner_id} if owner_id else {}
    rows = db.session.execute(
        text(f"""
            SELECT
                {owner_expr} AS owner_id,
                {abonent_id_expr} AS abonent_id,
                {uid_expr} AS account_uid,
                {account_expr} AS account_number,
                {fio_expr} AS fio,
                {address_expr} AS address
            FROM {a_table} a
            {where_sql}
            ORDER BY {owner_expr}, {account_expr}, {uid_expr}
        """),
        params,
    ).all()
    targets = []
    seen = set()
    for row in rows:
        data = dict(row._mapping)
        uid = _norm_text(data.get("account_uid"))
        owner = _norm_text(data.get("owner_id"))
        if not uid or (owner, uid) in seen:
            continue
        seen.add((owner, uid))
        targets.append({
            "owner_id": owner,
            "abonent_id": _norm_text(data.get("abonent_id")),
            "account_uid": uid,
            "account_number": _norm_text(data.get("account_number")),
            "identity": {
                "fio": _norm_text(data.get("fio")),
                "address": _norm_text(data.get("address")),
            },
        })
    return targets


def _latest_rows_by_owner_uid(model, owner_uids, uid_attr):
    if not owner_uids:
        return {}
    owners = sorted({_norm_text(owner) for owner, _uid in owner_uids if _norm_text(owner)})
    uids = sorted({_norm_text(uid) for _owner, uid in owner_uids if _norm_text(uid)})
    if not owners or not uids:
        return {}
    rows = (
        model.query
        .filter(model.owner_id.in_(owners))
        .filter(getattr(model, uid_attr).in_(uids))
        .order_by(model.updated_at.desc(), model.id.desc())
        .all()
    )
    result = {}
    wanted = set(owner_uids)
    for row in rows:
        key = (_norm_text(row.owner_id), _norm_text(getattr(row, uid_attr)))
        if key in wanted and key not in result:
            result[key] = row
    return result


def _safe_json_object(raw_json: str, invalid_reason: str):
    try:
        payload = json.loads(raw_json or "{}")
    except (TypeError, ValueError):
        return {}, invalid_reason
    if not isinstance(payload, dict):
        return {}, invalid_reason
    return payload, ""


def _audit_summary_payload(row: AbonentSummary | None, target: dict):
    if not row:
        return {
            "exists": False,
            "status": "missing",
            "reason": _cache_reason("summary", "missing", ""),
            "input_hash": "",
            "updated_at": None,
            "json": {},
            "json_error": "",
            "totals_error": "",
        }
    payload, json_error = _safe_json_object(row.summary_json, "SUMMARY_JSON_INVALID")
    status = _cache_status(row.summary_status or payload.get("summary_status") or payload.get("status"))
    reason = _cache_reason("summary", status, row.summary_reason or payload.get("summary_reason") or payload.get("reason") or json_error)
    return {
        "exists": True,
        "status": status,
        "reason": reason,
        "input_hash": _norm_text(row.input_hash or payload.get("input_hash")),
        "updated_at": _dt_json_or_none(row.updated_at),
        "json": payload,
        "json_error": json_error,
        "totals_error": _fresh_totals_validation_reason(payload) if status == "fresh" else "",
    }


def _audit_snapshot_payload(row: CardSnapshot | None):
    if not row:
        return {
            "exists": False,
            "status": "missing",
            "reason": _cache_reason("snapshot", "missing", ""),
            "input_hash": "",
            "updated_at": None,
            "json": {},
            "json_error": "",
        }
    payload, json_error = _safe_json_object(row.snapshot_json, "CARD_SNAPSHOT_JSON_INVALID")
    status = _cache_status(row.snapshot_status or payload.get("snapshot_status") or payload.get("summary_status") or payload.get("status"))
    reason = _cache_reason("snapshot", status, row.snapshot_reason or payload.get("snapshot_reason") or payload.get("summary_reason") or payload.get("reason") or json_error)
    return {
        "exists": True,
        "status": status,
        "reason": reason,
        "input_hash": _norm_text(row.input_hash or payload.get("input_hash")),
        "updated_at": _dt_json_or_none(row.updated_at),
        "json": payload,
        "json_error": json_error,
    }


def build_snapshot_summary_audit(owner_id: str = ""):
    targets = _audit_snapshot_summary_targets(owner_id)
    owner_uids = {
        (_norm_text(target.get("owner_id")), _norm_text(target.get("account_uid")))
        for target in targets
        if _norm_text(target.get("owner_id")) and _norm_text(target.get("account_uid"))
    }
    summaries = _latest_rows_by_owner_uid(AbonentSummary, owner_uids, "account_uid")
    snapshots = _latest_rows_by_owner_uid(CardSnapshot, owner_uids, "abonent_uid")
    counts = {
        "total_abonents": len(targets),
        "fresh_summary_count": 0,
        "error_count": 0,
        "missing_count": 0,
        "dirty_count": 0,
        "snapshot_missing_count": 0,
        "snapshot_dirty_error_count": 0,
        "hash_mismatch_count": 0,
    }
    items = []
    for target in targets:
        owner = _norm_text(target.get("owner_id"))
        uid = _norm_text(target.get("account_uid"))
        key = (owner, uid)
        summary = _audit_summary_payload(summaries.get(key), target)
        snapshot = _audit_snapshot_payload(snapshots.get(key))
        warnings = []

        if summary["status"] == "fresh":
            counts["fresh_summary_count"] += 1
            if summary.get("totals_error"):
                warnings.append(summary["totals_error"])
        elif summary["status"] in {"error", "invalid"}:
            counts["error_count"] += 1
        elif summary["status"] == "missing":
            counts["missing_count"] += 1
        elif summary["status"] == "dirty":
            counts["dirty_count"] += 1

        if snapshot["status"] == "missing" or not snapshot["exists"]:
            counts["snapshot_missing_count"] += 1
        if snapshot["status"] in {"dirty", "error", "invalid"}:
            counts["snapshot_dirty_error_count"] += 1

        if summary.get("json_error"):
            warnings.append(summary["json_error"])
        if snapshot.get("json_error"):
            warnings.append(snapshot["json_error"])
        if snapshot["status"] == "fresh" and (not snapshot["json"] or snapshot.get("json_error")):
            warnings.append("FRESH_SNAPSHOT_JSON_INVALID")

        hash_mismatch = summary["input_hash"] != snapshot["input_hash"]
        if hash_mismatch:
            counts["hash_mismatch_count"] += 1
            warnings.append("INPUT_HASH_CHANGED")
        if summary["status"] == "fresh" and not snapshot["exists"]:
            warnings.append("SUMMARY_FRESH_CARD_SNAPSHOT_MISSING")
        if summary["status"] == "fresh" and snapshot["status"] in {"dirty", "error", "invalid"}:
            warnings.append("SUMMARY_FRESH_SNAPSHOT_NOT_FRESH")
        if summary["status"] != "fresh" and not summary["reason"]:
            warnings.append("SUMMARY_REASON_MISSING")
        if snapshot["status"] != "fresh" and not snapshot["reason"]:
            warnings.append("SNAPSHOT_REASON_MISSING")

        items.append({
            "owner_id": owner,
            "account_number": _norm_text(target.get("account_number")),
            "account_uid": uid,
            "summary_status": summary["status"],
            "summary_reason": summary["reason"],
            "snapshot_status": snapshot["status"],
            "snapshot_reason": snapshot["reason"],
            "input_hash_summary": summary["input_hash"],
            "input_hash_snapshot": snapshot["input_hash"],
            "has_card_snapshot": bool(snapshot["exists"]),
            "has_abonent_summary": bool(summary["exists"]),
            "updated_at_summary": summary["updated_at"],
            "updated_at_snapshot": snapshot["updated_at"],
            "hash_mismatch": hash_mismatch,
            "warnings": sorted(set(warnings)),
        })

    return {
        "ok": True,
        "dry_run": True,
        "owner_id": _norm_text(owner_id),
        "counts": counts,
        "items": items,
    }


def _card_snapshot_payload(row: CardSnapshot):
    try:
        snapshot = json.loads(row.snapshot_json or "{}")
    except (TypeError, ValueError):
        snapshot = {}
    return {
        "owner_id": row.owner_id,
        "abonent_uid": row.abonent_uid or "",
        "abonent_id": row.abonent_id or "",
        "snapshot_status": row.snapshot_status or "missing",
        "snapshot_reason": row.snapshot_reason or "",
        "input_hash": row.input_hash or "",
        "ledger_version": row.ledger_version or "",
        "tariff_version": row.tariff_version or "",
        "rate_version": row.rate_version or "",
        "exclude_version": row.exclude_version or "",
        "links_version": row.links_version or "",
        "engine_version": row.engine_version or SNAPSHOT_ENGINE_VERSION,
        "computed_at": row.computed_at.isoformat() + "Z" if row.computed_at else None,
        "updated_at": row.updated_at.isoformat() + "Z" if row.updated_at else None,
        "snapshot": snapshot,
    }


def _snapshot_uid_from_payload(snapshot: dict):
    if not isinstance(snapshot, dict):
        return ""
    return _norm_text(
        snapshot.get("uid")
        or snapshot.get("account_uid")
        or snapshot.get("accountUid")
        or snapshot.get("abonent_uid")
        or snapshot.get("abonentUid")
    )


def _snapshot_target_for_owner(owner_id: str, uid: str):
    uid = _norm_text(uid)
    if not uid:
        return None
    return _owner_recalc_targets_by_uid(owner_id).get(uid)


@app.get("/api/audit/snapshot_summary")
def audit_snapshot_summary_endpoint():
    # Read-only diagnostics endpoint. Do not recalc, repair, rebuild, or mutate DB here.
    admin, err = _require_admin()
    if err:
        return err
    owner = _norm_text(request.args.get("owner") or request.args.get("owner_id") or "")
    try:
        return jsonify(build_snapshot_summary_audit(owner))
    except SQLAlchemyError as exc:
        db.session.rollback()
        app.logger.exception("snapshot_summary audit failed for admin=%s owner=%s", admin.id, owner)
        return jsonify(ok=False, error="snapshot_summary_audit_failed", details=str(exc)), 500


@app.get("/api/card_snapshot/<account_uid>")
def card_snapshot_get(account_uid: str):
    user, err = _require_user()
    if err:
        return err
    uid = _norm_text(account_uid)
    if not uid:
        return jsonify(ok=False, error="account_uid_required"), 400
    row = CardSnapshot.query.filter_by(owner_id=user.id, abonent_uid=uid).first()
    if not row:
        return jsonify(ok=True, snapshot_status="missing", snapshot_reason="SNAPSHOT_NOT_BUILT", snapshot=None)
    return jsonify(ok=True, **_card_snapshot_payload(row))


@app.post("/api/card_snapshot/<account_uid>")
def card_snapshot_put(account_uid: str):
    user, err = _require_user()
    if err:
        return err
    uid = _norm_text(account_uid)
    if not uid:
        return jsonify(ok=False, error="account_uid_required"), 400
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict):
        return jsonify(ok=False, error="snapshot_invalid"), 400
    snapshot = body.get("snapshot") if isinstance(body.get("snapshot"), dict) else body
    if not isinstance(snapshot, dict) or not snapshot:
        return jsonify(ok=False, error="snapshot_invalid"), 400
    snapshot_uid = _snapshot_uid_from_payload(snapshot) or _norm_text(body.get("abonent_uid") or body.get("account_uid"))
    if snapshot_uid and snapshot_uid != uid:
        return jsonify(ok=False, error="uid_mismatch"), 400
    target = _snapshot_target_for_owner(user.id, uid)
    if not target:
        return jsonify(ok=False, error="uid_not_found"), 404
    status = "fresh"
    reason = "OK"
    snapshot = dict(snapshot)
    snapshot["uid"] = uid
    snapshot["account_uid"] = uid
    snapshot["snapshot_status"] = status
    snapshot["summary_status"] = status
    snapshot["snapshot_reason"] = reason
    snapshot["summary_reason"] = reason
    row = CardSnapshot.query.filter_by(owner_id=user.id, abonent_uid=uid).first()
    if not row:
        row = CardSnapshot(owner_id=user.id, abonent_uid=uid)
        db.session.add(row)
    row.abonent_id = _norm_text(body.get("abonent_id") or snapshot.get("abonentId") or snapshot.get("abonent_id") or target.get("abonent_id"))
    row.snapshot_status = status
    row.snapshot_reason = reason
    row.input_hash = _norm_text(body.get("input_hash") or snapshot.get("input_hash"))
    row.ledger_version = _norm_text(body.get("ledger_version") or snapshot.get("ledgerVersion") or snapshot.get("ledger_version"))
    row.tariff_version = _norm_text(body.get("tariff_version") or snapshot.get("tariff_version"))
    row.rate_version = _norm_text(body.get("rate_version") or snapshot.get("rate_version"))
    row.exclude_version = _norm_text(body.get("exclude_version") or snapshot.get("exclude_version"))
    row.links_version = _norm_text(body.get("links_version") or snapshot.get("links_version"))
    row.engine_version = _norm_text(body.get("engine_version") or snapshot.get("engine_version")) or SNAPSHOT_ENGINE_VERSION
    row.computed_at = datetime.utcnow() if status == "fresh" else row.computed_at
    row.snapshot_json = json.dumps(snapshot, ensure_ascii=False, sort_keys=True)
    db.session.commit()
    return jsonify(ok=True, **_card_snapshot_payload(row))


@app.post("/api/recalc_lock/<account_uid>/begin")
def recalc_lock_begin(account_uid: str):
    user, err = _require_user()
    if err:
        return err
    uid = _norm_text(account_uid)
    if not uid:
        return jsonify(ok=False, error="account_uid_required"), 400
    now = datetime.utcnow()
    token = secrets.token_hex(16)
    existing = RecalcUidLock.query.filter_by(owner_id=user.id, abonent_uid=uid).first()
    if existing and existing.status == "running":
        age = (now - (existing.started_at or now)).total_seconds()
        if age <= RECALC_BATCH_RUNNING_TTL_SECONDS:
            return jsonify(ok=True, status="already_running", account_uid=uid)
        existing.status = "stale"
    if existing:
        existing.lock_token = token
        existing.status = "running"
        existing.started_at = now
    else:
        db.session.add(RecalcUidLock(owner_id=user.id, abonent_uid=uid, lock_token=token, status="running", started_at=now))
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        raced = RecalcUidLock.query.filter_by(owner_id=user.id, abonent_uid=uid).first()
        if raced and raced.status == "running":
            return jsonify(ok=True, status="already_running", account_uid=uid)
        app.logger.exception("recalc lock duplicate race recovery failed owner=%s uid=%s", user.id, uid)
        return jsonify(ok=False, error="recalc_lock_race_recovery_failed", account_uid=uid), 409
    return jsonify(ok=True, status="started", account_uid=uid, lock_token=token)


@app.post("/api/recalc_lock/<account_uid>/finish")
def recalc_lock_finish(account_uid: str):
    user, err = _require_user()
    if err:
        return err
    uid = _norm_text(account_uid)
    body = request.get_json(silent=True) or {}
    token = _norm_text(body.get("lock_token"))
    row = RecalcUidLock.query.filter_by(owner_id=user.id, abonent_uid=uid).first()
    if row and (not token or row.lock_token == token):
        row.status = "finished"
        db.session.commit()
    return jsonify(ok=True, status="finished", account_uid=uid)


@app.post("/api/auth/register")
def auth_register():
    data = request.get_json(silent=True) or {}
    email = _normalize_email(data.get("email"))
    password = str(data.get("password") or "")
    display_name = str(data.get("displayName") or "").strip()

    if not email:
        return _json_error("email_required", 400)
    if len(password) < 4:
        return _json_error("password_too_short", 400)
    if User.query.filter_by(email=email).first():
        return _json_error("email_exists", 409)

    role = "admin" if User.query.count() == 0 else "user"
    user = User(
        id="u_" + uuid.uuid4().hex[:12],
        email=email,
        password_hash=generate_password_hash(password),
        role=role,
        display_name=display_name,
        disabled=False,
        created_at=datetime.utcnow(),
        last_login=datetime.utcnow(),
    )
    db.session.add(user)
    db.session.commit()

    session.clear()
    session["user_id"] = user.id
    return jsonify(ok=True, user=_user_payload(user))


@app.post("/api/auth/login")
def auth_login():
    data = request.get_json(silent=True) or {}
    email = _normalize_email(data.get("email"))
    password = str(data.get("password") or "")

    if not email:
        return _json_error("email_required", 400)
    if not password:
        return _json_error("password_required", 400)

    user = User.query.filter_by(email=email).first()
    if not user:
        return _json_error("user_not_found", 404)
    if user.disabled:
        return _json_error("user_disabled", 403)
    if not check_password_hash(user.password_hash, password):
        return _json_error("invalid_password", 401)

    user.last_login = datetime.utcnow()
    db.session.commit()

    session.clear()
    session["user_id"] = user.id
    app.logger.info("[auth] login_ok user_id=%s email=%s role=%s", user.id, user.email, user.role)
    return jsonify(ok=True, user=_user_payload(user))


@app.get("/api/auth/me")
def auth_me():
    user = _current_user()
    if not user:
        return _json_error("not_authenticated", 401)
    app.logger.info("[auth] me_ok user_id=%s email=%s role=%s", user.id, user.email, user.role)
    return jsonify(ok=True, user=_user_payload(user))


@app.post("/api/auth/logout")
def auth_logout():
    user = _current_user()
    if user:
        app.logger.info("[auth] logout user_id=%s email=%s", user.id, user.email)
    session.clear()
    return jsonify(ok=True)


@app.get("/api/admin/users")
def admin_users_list():
    _admin, err = _require_admin()
    if err:
        return err
    users = User.query.order_by(User.created_at.asc(), User.email.asc()).all()
    return jsonify(ok=True, users=[_user_payload(u) for u in users])


@app.post("/api/admin/users")
def admin_users_create():
    _admin, err = _require_admin()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    email = _normalize_email(data.get("email"))
    password = str(data.get("password") or "")
    display_name = str(data.get("displayName") or "").strip()
    role = str(data.get("role") or "user").strip().lower()
    if role not in {"user", "admin", "moderator"}:
        role = "user"

    if not email:
        return _json_error("email_required", 400)
    if len(password) < 4:
        return _json_error("password_too_short", 400)
    if User.query.filter_by(email=email).first():
        return _json_error("email_exists", 409)

    user = User(
        id="u_" + uuid.uuid4().hex[:12],
        email=email,
        password_hash=generate_password_hash(password),
        role=role,
        display_name=display_name,
        disabled=False,
        created_at=datetime.utcnow(),
        last_login=None,
    )
    db.session.add(user)
    db.session.commit()
    return jsonify(ok=True, user=_user_payload(user))


@app.post("/api/admin/users/<user_id>/disable")
def admin_users_disable(user_id):
    admin, err = _require_admin()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    disabled = bool(data.get("disabled"))
    user = User.query.filter_by(id=user_id).first()
    if not user:
        return _json_error("user_not_found", 404)
    if admin.id == user.id and disabled:
        return _json_error("cannot_disable_self", 400)

    user.disabled = disabled
    db.session.commit()
    return jsonify(ok=True, user=_user_payload(user))


@app.post("/api/admin/users/<user_id>/password")
def admin_users_reset_password(user_id):
    _admin, err = _require_admin()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    new_password = str(data.get("password") or "")
    if len(new_password) < 4:
        return _json_error("password_too_short", 400)

    user = User.query.filter_by(id=user_id).first()
    if not user:
        return _json_error("user_not_found", 404)

    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    return jsonify(ok=True)


@app.delete("/api/admin/users/<user_id>")
def admin_users_delete(user_id):
    admin, err = _require_admin()
    if err:
        return err

    user = User.query.filter_by(id=user_id).first()
    if not user:
        return _json_error("user_not_found", 404)
    if user.id == admin.id:
        return _json_error("cannot_delete_self", 400)

    active_admins = User.query.filter_by(role="admin", disabled=False).count()
    if user.role == "admin" and active_admins <= 1:
        return _json_error("cannot_delete_last_admin", 400)

    db.session.delete(user)
    db.session.commit()
    return jsonify(ok=True)


@app.post("/api/import/payments/upload")
def import_payments_upload():
    user, err = _require_user()
    if err:
        return err

    owner, owner_err = _resolve_owner(request.form.get("owner"), allow_admin_override=True)
    if owner_err:
        return owner_err

    upload = request.files.get("file")
    if not upload:
        return _json_error("file_required", 400)

    file_bytes = upload.read()
    if not file_bytes:
        return _json_error("file_empty", 400)
    if len(file_bytes) > app.config["IMPORT_MAX_UPLOAD_BYTES"]:
        return jsonify(
            ok=False,
            error="file_too_large",
            details={"max_bytes": app.config["IMPORT_MAX_UPLOAD_BYTES"]},
        ), 400

    file_sha = hashlib.sha256(file_bytes).hexdigest()
    accounting_year = _norm_text(request.form.get("accounting_year"))
    display_name = _norm_text(request.form.get("display_name")) or _norm_text(upload.filename)

    existing_exact = ImportBatch.query.filter_by(owner_id=owner, file_sha256=file_sha).order_by(ImportBatch.id.asc()).first()
    existing_year = None
    if accounting_year:
        existing_year = ImportBatch.query.filter_by(owner_id=owner, accounting_year=accounting_year).order_by(ImportBatch.id.desc()).first()

    batch = ImportBatch(
        owner_id=owner,
        created_by_user_id=user.id,
        original_filename=_norm_text(upload.filename),
        file_sha256=file_sha,
        upload_blob=file_bytes,
        display_name=display_name,
        accounting_year=accounting_year,
        version_label=_norm_text(request.form.get("version_label")),
        duplicate_of_batch_id=existing_exact.id if existing_exact else None,
        supersedes_batch_id=int(request.form.get("supersedes_batch_id")) if _norm_text(request.form.get("supersedes_batch_id")) else None,
        overlaps_with_batch_id=existing_year.id if existing_year else None,
        is_exact_duplicate=bool(existing_exact),
        has_period_overlap=bool(existing_year),
        status="uploaded",
        uploaded_at=datetime.utcnow(),
        notes=(
            (_norm_text(request.form.get("notes")) + "\n") if _norm_text(request.form.get("notes")) else ""
        ) + f"upload_blob_policy: ttl_days={app.config['IMPORT_UPLOAD_BLOB_TTL_DAYS']}, max_bytes={app.config['IMPORT_MAX_UPLOAD_BYTES']}",
    )
    db.session.add(batch)
    db.session.commit()
    return jsonify(ok=True, batch=_batch_payload(batch))


@app.post("/api/import/payments/upload_rows")
def import_payments_upload_rows():
    user, err = _require_user()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    rows = data.get("rows")
    if not isinstance(rows, list) or not rows:
        return _json_error("rows_required", 400)

    owner, owner_err = _resolve_owner(data.get("owner"), allow_admin_override=True)
    if owner_err:
        return owner_err

    payload_dump = json.dumps(rows, ensure_ascii=False, sort_keys=True)
    payload_bytes = payload_dump.encode("utf-8")
    if len(payload_bytes) > app.config["IMPORT_MAX_UPLOAD_BYTES"]:
        return jsonify(
            ok=False,
            error="file_too_large",
            details={"max_bytes": app.config["IMPORT_MAX_UPLOAD_BYTES"]},
        ), 400

    file_sha = hashlib.sha256(payload_bytes).hexdigest()
    accounting_year = _norm_text(data.get("accounting_year"))
    display_name = _norm_text(data.get("display_name")) or "payments_rows.json"

    existing_exact = ImportBatch.query.filter_by(owner_id=owner, file_sha256=file_sha).order_by(ImportBatch.id.asc()).first()
    existing_year = None
    if accounting_year:
        existing_year = ImportBatch.query.filter_by(owner_id=owner, accounting_year=accounting_year).order_by(ImportBatch.id.desc()).first()

    batch = ImportBatch(
        owner_id=owner,
        created_by_user_id=user.id,
        original_filename="payments_rows.json",
        file_sha256=file_sha,
        upload_blob=b"",
        display_name=display_name,
        accounting_year=accounting_year,
        version_label=_norm_text(data.get("version_label")),
        duplicate_of_batch_id=existing_exact.id if existing_exact else None,
        supersedes_batch_id=int(data.get("supersedes_batch_id")) if _norm_text(data.get("supersedes_batch_id")) else None,
        overlaps_with_batch_id=existing_year.id if existing_year else None,
        is_exact_duplicate=bool(existing_exact),
        has_period_overlap=bool(existing_year),
        status="parsed",
        uploaded_at=datetime.utcnow(),
        started_at=datetime.utcnow(),
        file_name=_norm_text(data.get("file_name")) or "payments_rows.json",
        uploaded_by=user.id,
        notes=(
            (_norm_text(data.get("notes")) + "\n") if _norm_text(data.get("notes")) else ""
        ) + "source=upload_rows",
    )
    db.session.add(batch)
    db.session.flush()

    parsed_rows = []
    row_no = 0
    for idx, raw_row in enumerate(rows, start=1):
        row_no += 1
        src = raw_row if isinstance(raw_row, dict) else {}
        account_uid = _norm_text(src.get("account_uid"))
        abonent_id = _norm_text(src.get("abonent_id"))
        account_number = _norm_text(src.get("account_number"))
        payment_date = _norm_iso_date(src.get("payment_date"))
        payment_period = _norm_upload_payment_period(src.get("payment_period"))
        amount = _norm_amount(src.get("amount"))
        src_index_raw = src.get("source_index")
        try:
            source_index = int(src_index_raw) if src_index_raw is not None and _norm_text(src_index_raw) != "" else None
        except Exception:
            source_index = None

        charge_year, charge_month = _extract_year_month(payment_period)
        normalized_payload = {
            "account_uid": account_uid,
            "abonent_id": abonent_id,
            "account_number": account_number,
            "payment_date": payment_date,
            "payment_period": payment_period,
            "amount": amount,
            "source_index": source_index,
        }

        parsed_rows.append(ImportBatchRow(
            batch_id=batch.id,
            row_no=row_no,
            excel_sheet_name="upload_rows",
            excel_row_ref=f"row:{idx}",
            raw_payload_json=json.dumps(src, ensure_ascii=False),
            normalized_payload_json=json.dumps(normalized_payload, ensure_ascii=False),
            account_uid=account_uid,
            abonent_id=abonent_id,
            account_number=account_number,
            payment_date=payment_date or "",
            paid_date=payment_date or "",
            payment_period=payment_period or "",
            charge_year=charge_year,
            charge_month=charge_month,
            amount=amount or "",
            source_index=source_index,
            source_label=f"Платёж {source_index}" if source_index else "",
            fingerprint="",
            status="parsed",
            reason_code="",
            reason_text="",
            error_details_json="{}",
        ))

    db.session.bulk_save_objects(parsed_rows)
    batch.rows_total = row_no
    batch.started_at = datetime.utcnow()
    db.session.commit()
    return jsonify(ok=True, batch=_batch_payload(batch), parsed_rows=row_no)


@app.post("/api/import/<int:batch_id>/parse")
def import_payments_parse(batch_id):
    user, err = _require_user()
    if err:
        return err

    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)
    transition_err = _ensure_batch_transition(batch, "parse")
    if transition_err:
        return transition_err

    try:
        wb = load_workbook(io.BytesIO(batch.upload_blob), data_only=True)
    except Exception as ex:
        batch.status = "failed"
        batch.finished_at = datetime.utcnow()
        batch.error_message = "fingerprint_conflict"
        db.session.commit()
        return jsonify(ok=False, error="parse_failed", details=str(ex)), 400

    ImportBatchRow.query.filter_by(batch_id=batch.id).delete()
    row_no = 0
    parsed_rows = []

    missing_columns = set()
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        if not values:
            continue
        header_map = _parse_strict_header_map(values[0])
        missing = sorted(IMPORT_REQUIRED_COLUMNS.difference(set(header_map.keys())))
        if missing:
            missing_columns.update(missing)
            continue
        for idx in range(2, len(values) + 1):
            line = values[idx - 1]
            if all(_norm_text(v) == "" for v in line if v is not None):
                continue
            row_no += 1
            account_uid = _norm_text(_find_cell(header_map, line, "account_uid"))
            abonent_id = _norm_text(_find_cell(header_map, line, "abonent_id"))
            account_number = _norm_text(_find_cell(header_map, line, "account_number"))
            payment_date = _norm_date(_find_cell(header_map, line, "payment_date"))
            payment_period = _norm_period(_find_cell(header_map, line, "payment_period"))
            amount = _norm_amount(_find_cell(header_map, line, "amount"))
            src_index_raw = _find_cell(header_map, line, "source_index")
            try:
                source_index = int(src_index_raw) if src_index_raw is not None and _norm_text(src_index_raw) != "" else None
            except Exception:
                source_index = None

            charge_year, charge_month = _extract_year_month(payment_period)
            normalized_payload = {
                "account_uid": account_uid,
                "abonent_id": abonent_id,
                "account_number": account_number,
                "payment_date": payment_date,
                "payment_period": payment_period,
                "amount": amount,
                "source_index": source_index,
            }
            parsed_rows.append(ImportBatchRow(
                batch_id=batch.id,
                row_no=row_no,
                excel_sheet_name=ws.title,
                excel_row_ref=f"{ws.title}:{idx}",
                raw_payload_json=json.dumps({"values": ["" if v is None else str(v) for v in line]}, ensure_ascii=False),
                normalized_payload_json=json.dumps(normalized_payload, ensure_ascii=False),
                account_uid=account_uid,
                abonent_id=abonent_id,
                account_number=account_number,
                payment_date=payment_date or "",
                paid_date=payment_date or "",
                payment_period=payment_period or "",
                charge_year=charge_year,
                charge_month=charge_month,
                amount=amount or "",
                source_index=source_index,
                source_label=f"Платёж {source_index}" if source_index else "",
                fingerprint="",
                status="parsed",
                reason_code="",
                reason_text="",
                error_details_json="{}",
            ))

    if missing_columns:
        batch.status = "failed"
        batch.finished_at = datetime.utcnow()
        db.session.commit()
        return jsonify(
            ok=False,
            error="template_header_mismatch",
            details={"missing_required_columns": sorted(missing_columns)},
        ), 400

    db.session.bulk_save_objects(parsed_rows)
    batch.rows_total = row_no
    batch.status = "parsed"
    batch.started_at = datetime.utcnow()
    db.session.commit()
    return jsonify(ok=True, batch=_batch_payload(batch), parsed_rows=row_no, summary={"missing_required_columns": []})


@app.post("/api/import/<int:batch_id>/validate")
def import_payments_validate(batch_id):
    user, err = _require_user()
    if err:
        return err

    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)
    transition_err = _ensure_batch_transition(batch, "validate")
    if transition_err:
        return transition_err

    rows = ImportBatchRow.query.filter_by(batch_id=batch.id).order_by(ImportBatchRow.row_no.asc()).all()
    if not rows:
        return _json_error("rows_not_found", 400)

    seen = set()
    seen_uid_date = {}
    sources_map = _load_owner_sources(batch.owner_id)
    valid = invalid = duplicate = 0

    for r in rows:
        details = {}
        if not r.account_uid:
            r.status = "invalid"
            r.reason_code = "ACCOUNT_UID_REQUIRED"
            r.reason_text = "Платежи могут учитываться только при наличии UID"
            details["field"] = "account_uid"
            details["recommendation"] = "Заполните account_uid существующим UID абонента."
            invalid += 1
        elif not r.account_number:
            r.status = "invalid"
            r.reason_code = "ACCOUNT_NUMBER_REQUIRED"
            r.reason_text = "Для записи оплаты в ledger нужен лицевой счёт"
            details["field"] = "account_number"
            details["recommendation"] = "Заполните лицевой счёт (account_number)."
            invalid += 1
        elif not r.payment_date:
            r.status = "invalid"
            r.reason_code = "PAYMENT_DATE_INVALID"
            r.reason_text = "Некорректная дата платежа"
            details["field"] = "payment_date"
            details["recommendation"] = "Проверьте дату платежа, ожидается YYYY-MM-DD."
            invalid += 1
        elif not r.payment_period:
            r.status = "invalid"
            r.reason_code = "PAYMENT_PERIOD_INVALID" if _raw_upload_payment_period_present(r) else "PAYMENT_PERIOD_REQUIRED"
            r.reason_text = "Не найден год/период платежа. Импорт строки остановлен, чтобы не записать платёж не в тот год."
            details["field"] = "payment_period"
            details["recommendation"] = "Проверьте период платежа, ожидается YYYY-MM."
            invalid += 1
        elif not r.amount:
            r.status = "invalid"
            r.reason_code = "AMOUNT_INVALID"
            r.reason_text = "Некорректная сумма"
            details["field"] = "amount"
            details["recommendation"] = "Проверьте сумму, ожидается число > 0."
            invalid += 1
        elif r.source_index is None:
            r.status = "invalid"
            r.reason_code = "SOURCE_INDEX_INVALID"
            r.reason_text = "source_index обязателен"
            details["field"] = "source_index"
            details["recommendation"] = "Укажите индекс источника платежа."
            invalid += 1
        elif r.source_index not in sources_map:
            r.status = "invalid"
            r.reason_code = "SOURCE_INDEX_INVALID"
            r.reason_text = "source_index отсутствует в справочнике payment_sources"
            details["field"] = "source_index"
            details["recommendation"] = "Используйте source_index из справочника источников оплаты."
            invalid += 1
        else:
            lookup = _find_owner_accounts(batch.owner_id, r.account_uid, r.account_number)
            matches = lookup["matches"]
            if not matches and lookup["uid_found"]:
                r.status = "invalid"
                r.reason_code = "UID_LS_MISMATCH"
                r.reason_text = "UID найден, но лицевой счёт не совпадает"
                details["field"] = "account_number"
                details["recommendation"] = "Проверьте соответствие UID и лицевого счёта."
                invalid += 1
            elif not matches:
                r.status = "invalid"
                r.reason_code = "ACCOUNT_NOT_FOUND"
                r.reason_text = "account_uid не найден у текущего owner"
                details["field"] = "account_uid"
                details["recommendation"] = "Проверьте UID/ЛС и загрузите корректные данные абонента."
                invalid += 1
            elif len(matches) > 1:
                r.status = "invalid"
                r.reason_code = "AMBIGUOUS_ACCOUNT_MATCH"
                r.reason_text = "Найдено несколько совпадений account_uid/account_number"
                details["field"] = "account_uid"
                details["recommendation"] = "Уточните ЛС/UID для однозначного сопоставления."
                invalid += 1
            else:
                r.source_label = sources_map.get(r.source_index, r.source_label)
                try:
                    normalized_uid = normalize_uid(r.account_uid)
                    normalized_account_number = normalize_account_number(r.account_number)
                    normalized_paid_date = normalize_paid_date(r.payment_date)
                    normalized_period = normalize_payment_period(r.payment_period)
                    normalized_amount = normalize_amount(r.amount)
                    normalized_source_index = normalize_source_index(r.source_index)
                except ValueError as ex:
                    r.status = "invalid"
                    r.reason_code = str(ex)
                    r.reason_text = "Некорректные данные строки для fingerprint"
                    details["field"] = "fingerprint"
                    details["recommendation"] = "Исправьте обязательные поля и повторите валидацию."
                    invalid += 1
                    r.error_details_json = json.dumps(details, ensure_ascii=False)
                    continue

                r.account_uid = normalized_uid
                r.account_number = normalized_account_number
                r.payment_date = normalized_paid_date
                r.paid_date = normalized_paid_date
                r.payment_period = normalized_period
                r.amount = normalized_amount
                r.source_index = normalized_source_index
                r.fingerprint = build_payment_fingerprint(
                    batch.owner_id,
                    r.account_uid,
                    r.account_number,
                    r.paid_date,
                    r.amount,
                    r.source_index,
                    r.payment_period,
                )
                uid_date_key = f"{r.account_uid}|{r.paid_date}"
                if uid_date_key in seen_uid_date and seen_uid_date[uid_date_key] != r.amount:
                    r.status = "conflict"
                    r.reason_code = "CONFLICT"
                    r.reason_text = "Найден платёж с тем же UID+paid_date, но другой суммой"
                    duplicate += 1
                elif r.fingerprint in seen:
                    r.status = "duplicate"
                    r.reason_code = "DUPLICATE"
                    r.reason_text = "Дубликат в текущем батче"
                    duplicate += 1
                else:
                    uid_key_part = r.account_uid
                    key = f"payments_{uid_key_part}"
                    kv = KVStore.query.filter_by(owner=batch.owner_id, k=key).first()
                    try:
                        ledger = _load_existing_payment_ledger_or_raise(kv)
                    except LedgerJsonInvalidError:
                        r.status = "invalid"
                        r.reason_code = "LEDGER_JSON_INVALID"
                        r.reason_text = "Данные платежей повреждены. Расчёт/импорт остановлен, чтобы не потерять историю платежей."
                        invalid += 1
                        continue

                    classification = _classify_import_payment(r.account_uid, r.paid_date, r.amount, r.fingerprint, ledger)
                    if classification == "NEW_PAYMENT":
                        seen.add(r.fingerprint)
                        seen_uid_date[uid_date_key] = r.amount
                        r.status = "ready"
                        r.reason_code = "NEW_PAYMENT"
                        r.reason_text = ""
                        valid += 1
                    elif classification == "DUPLICATE":
                        r.status = "duplicate"
                        r.reason_code = "DUPLICATE"
                        r.reason_text = "Платёж уже существует"
                        duplicate += 1
                    else:
                        r.status = "conflict"
                        r.reason_code = "CONFLICT"
                        r.reason_text = "Найден платёж с тем же UID+paid_date, но другой суммой"
                        duplicate += 1
        r.error_details_json = json.dumps(details, ensure_ascii=False)

    batch.rows_valid = valid
    batch.rows_invalid = invalid
    batch.rows_duplicate = duplicate
    batch.status = "ready_to_apply" if invalid == 0 else "validated"
    db.session.commit()
    return jsonify(ok=True, batch=_batch_payload(batch))


@app.get("/api/import/<int:batch_id>/rows")
def import_payments_rows(batch_id):
    user, err = _require_user()
    if err:
        return err
    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)

    q = ImportBatchRow.query.filter_by(batch_id=batch.id)
    status = _norm_text(request.args.get("status"))
    if status:
        q = q.filter_by(status=status)
    rows = q.order_by(ImportBatchRow.row_no.asc()).all()
    return jsonify(ok=True, rows=[_row_payload(r) for r in rows])


@app.post("/api/import/<int:batch_id>/apply")
def import_payments_apply(batch_id):
    user, err = _require_user()
    if err:
        return err
    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)
    transition_err = _ensure_batch_transition(batch, "apply")
    if transition_err:
        return transition_err

    if batch.status == "failed":
        return jsonify(
            ok=False,
            error="batch_failed_restart_required",
            details={
                "message": "Батч завершился с ошибкой. Требуется повторная загрузка."
            }
        ), 400

    if batch.rows_invalid > 0:
        return jsonify(
            ok=False,
            error="invalid_rows_present",
            details={
                "rows_invalid": batch.rows_invalid,
                "message": "Импорт невозможен: в батче есть строки с ошибками. Исправьте их и повторите валидацию.",
            },
        ), 400

    rows = ImportBatchRow.query.filter_by(batch_id=batch.id).order_by(ImportBatchRow.row_no.asc()).all()
    applicable_statuses = {"ready", "duplicate"}
    not_ready = [r for r in rows if r.status not in applicable_statuses]
    if not_ready:
        return jsonify(
            ok=False,
            error="non_ready_rows_present",
            details={"count": len(not_ready)},
        ), 400
    applied_count = skipped_count = duplicate_count = conflict_count = failed_count = 0
    affected_uids = set()
    sources_map = _load_owner_sources(batch.owner_id)
    current_row_id = None
    try:
        batch.status = "applying"
        db.session.flush()

        for r in rows:
            current_row_id = r.id
            if r.status != "ready":
                action = r.reason_code or "SKIPPED"
                if r.reason_code == "DUPLICATE":
                    duplicate_count += 1
                elif r.reason_code == "CONFLICT":
                    conflict_count += 1
                else:
                    skipped_count += 1
                db.session.add(PaymentAuditLog(
                    owner_id=batch.owner_id,
                    batch_id=batch.id,
                    row_id=r.id,
                    action=action,
                    status="SKIPPED",
                    details_json=json.dumps({"account_uid": r.account_uid, "payment_date": r.paid_date or r.payment_date, "amount": r.amount, "source_index": r.source_index, "result": "SKIPPED", "reason_code": r.reason_code, "reason_text": r.reason_text, "fingerprint": r.fingerprint}, ensure_ascii=False),
                ))
                continue

            normalized_uid = normalize_uid(r.account_uid)
            normalized_account_number = normalize_account_number(r.account_number)
            normalized_paid_date = normalize_paid_date(r.paid_date or r.payment_date)
            normalized_period = normalize_payment_period(r.payment_period)
            normalized_amount = normalize_amount(r.amount)
            normalized_source_index = normalize_source_index(r.source_index)
            fingerprint = payment_fingerprint(normalized_uid, normalized_paid_date, normalized_amount, normalized_source_index)

            uid_key_part = normalized_uid
            key = f"payments_{uid_key_part}"
            existing_fingerprint = ImportAppliedFingerprint.query.filter_by(
                owner_id=batch.owner_id,
                import_type="payments",
                fingerprint=fingerprint,
            ).first()
            if existing_fingerprint:
                r.status = "duplicate"
                r.reason_code = "DUPLICATE"
                r.reason_text = "Платёж уже применён ранее"
                r.matched_payment_id = existing_fingerprint.payment_id or ""
                duplicate_count += 1
                db.session.add(PaymentAuditLog(
                    owner_id=batch.owner_id,
                    batch_id=batch.id,
                    row_id=r.id,
                    action="DUPLICATE",
                    status="SKIPPED",
                    details_json=json.dumps({"account_uid": normalized_uid, "payment_date": normalized_paid_date, "amount": normalized_amount, "source_index": normalized_source_index, "result": "DUPLICATE", "reason_code": r.reason_code, "fingerprint": fingerprint}, ensure_ascii=False),
                ))
                continue

            fingerprint_row = ImportAppliedFingerprint(
                owner_id=batch.owner_id,
                import_type="payments",
                fingerprint=fingerprint,
                account_uid=normalized_uid,
                account_number=normalized_account_number,
                payment_period=normalized_period,
                paid_date=datetime.strptime(normalized_paid_date, "%Y-%m-%d").date(),
                amount=Decimal(normalized_amount),
                source_index=normalized_source_index,
                batch_id=batch.id,
            )
            db.session.add(fingerprint_row)

            kv = KVStore.query.filter_by(owner=batch.owner_id, k=key).with_for_update().first()
            ledger = []
            source_kv = kv
            try:
                ledger = _load_existing_payment_ledger_or_raise(source_kv)
            except LedgerJsonInvalidError:
                raise LedgerJsonInvalidError("LEDGER_JSON_INVALID")

            max_id = 0
            for x in ledger:
                try:
                    max_id = max(max_id, int(x.get("id") or 0))
                except Exception:
                    continue
            next_id = max_id + 1
            yy, mm = _extract_year_month(normalized_period)
            ledger_item = {
                "id": next_id,
                "uid": normalized_uid,
                "year": yy,
                "month": mm,
                "accrued": 0,
                "paid": float(normalized_amount),
                "paid_date": to_ledger_paid_date(normalized_paid_date),
                "source": sources_map.get(normalized_source_index) or r.source_label or f"Платёж {normalized_source_index}",
                "payment_period": normalized_period,
                "fingerprint": fingerprint,
            }
            ledger.append(ledger_item)
            if kv:
                kv.v = json.dumps(ledger, ensure_ascii=False)
            else:
                db.session.add(KVStore(owner=batch.owner_id, k=key, v=json.dumps(ledger, ensure_ascii=False)))

            payment_id = f"{normalized_account_number}:{next_id}"
            fingerprint_row.payment_id = payment_id
            r.account_uid = normalized_uid
            r.account_number = normalized_account_number
            r.payment_date = normalized_paid_date
            r.paid_date = normalized_paid_date
            r.payment_period = normalized_period
            r.amount = normalized_amount
            r.source_index = normalized_source_index
            r.fingerprint = fingerprint
            r.matched_payment_id = payment_id
            r.applied_at = datetime.utcnow()
            r.status = "applied"
            r.reason_code = ""
            r.reason_text = ""
            db.session.add(PaymentAuditLog(
                owner_id=batch.owner_id,
                batch_id=batch.id,
                row_id=r.id,
                action="APPLIED",
                status="APPLIED",
                details_json=json.dumps({"account_uid": normalized_uid, "payment_date": normalized_paid_date, "amount": normalized_amount, "source_index": normalized_source_index, "result": "APPLIED", "payment_id": payment_id, "fingerprint": fingerprint}, ensure_ascii=False),
            ))
            applied_count += 1
            affected_uids.add(normalized_uid)

        batch.rows_applied = applied_count
        batch.rows_skipped = duplicate_count + conflict_count + skipped_count
        batch.error_message = ""
        batch.status = "applied"
        batch.finished_at = datetime.utcnow()
        if batch.uploaded_at and app.config["IMPORT_UPLOAD_BLOB_TTL_DAYS"] <= 0:
            batch.upload_blob = b""
        db.session.commit()
    except LedgerJsonInvalidError as ex:
        db.session.rollback()
        batch = ImportBatch.query.filter_by(id=batch.id).first()
        if batch:
            batch.status = "failed"
            batch.finished_at = datetime.utcnow()
            batch.error_message = "LEDGER_JSON_INVALID"
            db.session.execute(
                text("UPDATE import_rows SET status = 'failed', reason_code = 'LEDGER_JSON_INVALID', reason_text = :reason_text WHERE batch_id = :batch_id"),
                {"batch_id": batch.id, "reason_text": "Данные платежей повреждены. Расчёт/импорт остановлен, чтобы не потерять историю платежей."},
            )
            db.session.add(PaymentAuditLog(
                owner_id=batch.owner_id,
                batch_id=batch.id,
                action="FAILED",
                status="FAILED",
                details_json=json.dumps({"error": "LEDGER_JSON_INVALID", "row_id": current_row_id}, ensure_ascii=False),
            ))
            db.session.commit()
        failed_count = 1
        return jsonify(ok=False, error="apply_failed", details="LEDGER_JSON_INVALID"), 500
    except IntegrityError as e:
        db.session.rollback()
        batch = ImportBatch.query.filter_by(id=batch.id).first()
        if batch:
            batch.status = "failed"
            batch.finished_at = datetime.utcnow()
            batch.error_message = "fingerprint_conflict"
            db.session.execute(
                text("UPDATE import_rows SET status = 'failed' WHERE batch_id = :batch_id"),
                {"batch_id": batch.id},
            )
            db.session.add(PaymentAuditLog(
                owner_id=batch.owner_id,
                batch_id=batch.id,
                action="FAILED",
                status="FAILED",
                details_json=json.dumps({"error": str(e)}, ensure_ascii=False),
            ))
            db.session.commit()
        failed_count = 1
        return jsonify(ok=False, error="apply_failed", details="fingerprint_conflict"), 500
    except Exception as ex:
        db.session.rollback()
        batch = ImportBatch.query.filter_by(id=batch.id).first()
        if batch:
            batch.status = "failed"
            batch.finished_at = datetime.utcnow()
            batch.error_message = str(ex)
            db.session.execute(
                text("UPDATE import_rows SET status = 'failed' WHERE batch_id = :batch_id"),
                {"batch_id": batch.id},
            )
            db.session.add(PaymentAuditLog(
                owner_id=batch.owner_id,
                batch_id=batch.id,
                action="FAILED",
                status="FAILED",
                details_json=json.dumps({"error": str(ex), "row_id": current_row_id}, ensure_ascii=False),
            ))
            db.session.commit()
        failed_count = 1
        return jsonify(ok=False, error="apply_failed", details=str(ex)), 500

    summary = {
        "applied_count": applied_count,
        "skipped_count": skipped_count,
        "duplicate_count": duplicate_count,
        "conflict_count": conflict_count,
        "failed_count": failed_count,
        "affected_uids": sorted(affected_uids),
    }
    return jsonify(ok=True, batch=_batch_payload(batch), summary=summary)


@app.get("/api/import/payments/batches")
def import_payments_batches():
    user, err = _require_user()
    if err:
        return err
    owner = request.args.get("owner") if user.role == "admin" else user.id
    q = ImportBatch.query
    if owner:
        q = q.filter_by(owner_id=owner)
    items = q.order_by(ImportBatch.uploaded_at.desc()).limit(200).all()
    return jsonify(ok=True, batches=[_batch_payload(x) for x in items])


@app.get("/api/import/<int:batch_id>/summary")
def import_payments_summary(batch_id):
    user, err = _require_user()
    if err:
        return err
    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)

    return jsonify(
        ok=True,
        summary={
            "batch_id": batch.id,
            "status": batch.status,
            "file_name": batch.file_name or batch.original_filename,
            "rows_total": batch.rows_total,
            "rows_valid": batch.rows_valid,
            "rows_invalid": batch.rows_invalid,
            "rows_applied": batch.rows_applied,
            "rows_skipped": batch.rows_skipped,
            "started_at": batch.started_at.isoformat() + "Z" if batch.started_at else None,
            "finished_at": batch.finished_at.isoformat() + "Z" if batch.finished_at else None,
        },
    )


@app.get("/api/import/<int:batch_id>/errors")
def import_payments_errors(batch_id):
    user, err = _require_user()
    if err:
        return err
    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)

    bad = ImportBatchRow.query.filter(
        ImportBatchRow.batch_id == batch.id,
        ImportBatchRow.status.in_(["invalid", "failed", "duplicate", "conflict"]),
    ).order_by(ImportBatchRow.row_no.asc()).all()
    by_reason = {}
    for r in bad:
        by_reason[r.reason_code or "unknown"] = by_reason.get(r.reason_code or "unknown", 0) + 1
    return jsonify(
        ok=True,
        errors=[_row_human_error_payload(r) for r in bad],
        summary={"total": len(bad), "by_reason": by_reason},
    )


@app.get("/api/import/<int:batch_id>/errors/export")
def import_payments_errors_export(batch_id):
    user, err = _require_user()
    if err:
        return err
    batch = ImportBatch.query.filter_by(id=batch_id).first()
    if not batch:
        return _json_error("batch_not_found", 404)
    if user.role != "admin" and batch.owner_id != user.id:
        return _json_error("forbidden", 403)

    bad = ImportBatchRow.query.filter(
        ImportBatchRow.batch_id == batch.id,
        ImportBatchRow.status.in_(["invalid", "failed", "duplicate", "conflict"]),
    ).order_by(ImportBatchRow.row_no.asc()).all()

    fmt = _norm_text(request.args.get("format")).lower() or "csv"
    headers = [
        "excel_row_ref",
        "excel_sheet_name",
        "account_uid",
        "account_number",
        "payment_date",
        "payment_period",
        "amount",
        "source_index",
        "source_label",
        "reason_code",
        "reason_text",
        "recommendation",
    ]

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "errors"
        ws.append(headers)
        for r in bad:
            ws.append([
                r.excel_row_ref,
                r.excel_sheet_name,
                r.account_uid,
                r.account_number,
                r.payment_date,
                r.payment_period,
                r.amount,
                r.source_index,
                r.source_label,
                r.reason_code,
                r.reason_text,
                json.loads(r.error_details_json or "{}").get("recommendation", ""),
            ])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return Response(
            bio.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=import_errors_{batch.id}.xlsx"},
        )

    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow(headers)
    for r in bad:
        writer.writerow([
            r.excel_row_ref,
            r.excel_sheet_name,
            r.account_uid,
            r.account_number,
            r.payment_date,
            r.payment_period,
            r.amount,
            r.source_index,
            r.source_label,
            r.reason_code,
            r.reason_text,
            json.loads(r.error_details_json or "{}").get("recommendation", ""),
        ])
    return Response(
        sio.getvalue().encode("utf-8-sig"),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=import_errors_{batch.id}.csv"},
    )


@app.get("/api/store_keys")
def store_keys():
    requested_owner = request.args.get("owner") or request.args.get("owner_id")
    owner, err = _resolve_owner(requested_owner, allow_admin_override=True)
    if err:
        return err

    rows_owner = db.session.execute(
        text("SELECT k FROM kv_store WHERE owner=:owner ORDER BY k"),
        {"owner": owner},
    ).all()
    rows_global = db.session.execute(
        text(
            "SELECT k FROM kv_store WHERE owner=:owner "
            "AND k IN ('refinancing_rates_normal_v1','refinancing_rates_moratorium_v1') ORDER BY k"
        ),
        {"owner": GLOBAL_OWNER},
    ).all()
    keys = sorted({r[0] for r in rows_owner}.union({r[0] for r in rows_global}))
    _sync_log("list_keys", owner, count=len(keys))
    return jsonify(ok=True, keys=keys, owner=owner)


@app.get("/api/store")
def store_get():
    requested_owner = request.args.get("owner") or request.args.get("owner_id")
    owner, err = _resolve_owner(requested_owner, allow_admin_override=True)
    if err:
        return err

    key = (request.args.get("key") or "").strip()
    if not key:
        return _json_error("key_required", 400)

    owner_eff = _effective_owner_for_key(owner, key)
    row = KVStore.query.filter_by(owner=owner_eff, k=key).first()
    if not row:
        _sync_log("load", owner_eff, key=key, status="not_found")
        return jsonify(ok=False, error="not_found", value=None), 404
    _sync_log("load", owner_eff, key=key, size=len(row.v or ""), status="ok")
    return jsonify(ok=True, value=row.v, owner=owner_eff)


@app.post("/api/store")
def store_set():
    data = request.get_json(silent=True) or {}
    user, user_err = _require_user()
    if user_err:
        return user_err
    owner, err = _resolve_owner(data.get("owner"), allow_admin_override=True)
    if err:
        return err

    key = (data.get("key") or "").strip()
    value = data.get("value")
    if not key:
        return _json_error("key_required", 400)
    if value is None:
        value = ""
    if not isinstance(value, str):
        return _json_error("value_must_be_string", 400)
    access_err, reason = _require_write_access_for_key(owner, key)
    if access_err:
        _log_store_forbidden(user, owner, key, reason or "forbidden")
        return access_err

    owner_eff = _effective_owner_for_key(owner, key)
    row = KVStore.query.filter_by(owner=owner_eff, k=key).first()
    if row:
        row.v = value
    else:
        db.session.add(KVStore(owner=owner_eff, k=key, v=value))
    db.session.commit()
    _sync_log("save", owner_eff, key=key, size=len(value or ""), status="ok")
    return jsonify(ok=True, owner=owner_eff)


@app.delete("/api/store")
def store_delete():
    data = request.get_json(silent=True) or {}
    user, user_err = _require_user()
    if user_err:
        return user_err
    owner, err = _resolve_owner(data.get("owner"), allow_admin_override=True)
    if err:
        return err

    key = (data.get("key") or "").strip()
    if not key:
        return _json_error("key_required", 400)
    access_err, reason = _require_write_access_for_key(owner, key)
    if access_err:
        _log_store_forbidden(user, owner, key, reason or "forbidden")
        return access_err

    owner_eff = _effective_owner_for_key(owner, key)
    row = KVStore.query.filter_by(owner=owner_eff, k=key).first()
    if not row:
        _sync_log("delete", owner_eff, key=key, status="not_found")
        return jsonify(ok=True, deleted=False)

    db.session.delete(row)
    db.session.commit()
    _sync_log("delete", owner_eff, key=key, status="ok")
    return jsonify(ok=True, deleted=True)


@app.get("/api/store_dump")
def store_dump():
    requested_owner = request.args.get("owner") or request.args.get("owner_id")
    owner, err = _resolve_owner(requested_owner, allow_admin_override=True)
    if err:
        return err

    rows_owner = db.session.execute(
        text("SELECT k, v FROM kv_store WHERE owner=:owner ORDER BY k"),
        {"owner": owner},
    ).all()
    rows_global = db.session.execute(
        text(
            "SELECT k, v FROM kv_store WHERE owner=:owner "
            "AND k IN ('refinancing_rates_normal_v1','refinancing_rates_moratorium_v1') ORDER BY k"
        ),
        {"owner": GLOBAL_OWNER},
    ).all()
    data = {r[0]: r[1] for r in rows_owner}
    for r in rows_global:
        data[r[0]] = r[1]
    _sync_log("dump", owner, keys=len(data), status="ok")
    return jsonify(ok=True, owner=owner, data=data)


@app.get("/api/admin/session_debug")
def admin_session_debug():
    user = _current_user()
    return jsonify(ok=True, session_user_id=session.get("user_id"), user=_user_payload(user) if user else None)


@app.post("/api/ref_rates/error_report")
def ref_rates_error_report():
    user, err = _require_user()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    message = str(data.get("message") or "").strip()
    if not message:
        return _json_error("message_required", 400)
    if len(message) > 2000:
        return _json_error("message_too_long", 400)

    owner, owner_err = _resolve_owner(data.get("owner"))
    if owner_err:
        return owner_err

    app.logger.warning(
        "[ref_rates][error_report] owner=%s user_id=%s role=%s message=%s",
        owner,
        user.id,
        user.role,
        message,
    )
    return jsonify(ok=True, owner=owner)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
